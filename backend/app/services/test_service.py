"""Testlar — tuzish, ishlash, avtomatik tekshiruv (TST-01…TST-05).

Beshta qoida modulni belgilaydi:

1. **Toʻgʻri javob oʻquvchiga hech qachon yuborilmaydi.** Oʻquvchi
   uchun variantlar `is_correct` USTUNISIZ soʻraladi — sxemada kesib
   tashlashga tayanilmaydi. Bitta unutilgan `response_model` butun
   testni ochib berardi (X-5).

2. **Baholash serverda.** Ball frontenddan qabul qilinmaydi; javoblar
   yuborilganda server oʻzi hisoblaydi (TST-04).

3. **Urinishlar soni serverda cheklanadi** (TST-03). Frontenddagi
   tugmani yashirish himoya emas.

4. **Test faqat oʻz sinfiga.** Oʻquvchi va ota-ona boshqa sinfning
   testini koʻra olmaydi — kesim soʻrov darajasida (X-1).

5. **Hech narsa oʻchirilmaydi.** Savol ham, test ham arxivlanadi:
   oʻtgan urinishlar oʻsha savollarga bogʻlangan.
"""

import io
import random
import uuid
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import (
    ConflictError,
    NotFoundError,
    PermissionDeniedError,
    ValidationError,
)
from app.core.timeutil import utcnow
from app.models import (
    AuditAction,
    QuestionKind,
    SchoolClass,
    Student,
    Subject,
    Test,
    TestAnswer,
    TestAttempt,
    TestOption,
    TestQuestion,
    TestStatus,
)
from app.services import audit_service
from app.services.access import (
    CurrentUser,
    assert_can_view_student,
    assert_student_self,
)

MAX_OPTIONS = 8
MAX_QUESTIONS = 100


@dataclass(frozen=True, slots=True)
class TestRow:
    id: uuid.UUID
    class_id: uuid.UUID
    class_name: str
    subject_id: uuid.UUID
    subject_name: str
    title: str
    description: str
    status: str
    duration_minutes: int
    attempts_allowed: int
    shuffle: bool
    opens_at: datetime
    closes_at: datetime
    question_count: int
    max_score: int
    #: TST-05: nechta oʻquvchi topshirdi va oʻrtacha foiz.
    submitted_count: int
    total_students: int
    average_percent: float | None


@dataclass(frozen=True, slots=True)
class OptionRow:
    id: uuid.UUID
    text: str
    #: Oʻquvchiga `None` — toʻgʻri javob yuborilmaydi (1-qoida).
    is_correct: bool | None


@dataclass(frozen=True, slots=True)
class QuestionRow:
    id: uuid.UUID
    position: int
    text: str
    kind: str
    points: int
    options: list[OptionRow]


@dataclass(frozen=True, slots=True)
class AttemptRow:
    id: uuid.UUID
    test_id: uuid.UUID
    student_id: uuid.UUID
    full_name: str
    attempt_no: int
    started_at: datetime
    submitted_at: datetime | None
    score: int | None
    max_score: int
    percent: float | None


# ─────────────────────────── Ruxsat ───────────────────────────


async def _load_for_teacher(session: AsyncSession, user: CurrentUser, test_id: uuid.UUID) -> Test:
    test = await session.get(Test, test_id)
    if test is None or test.is_archived:
        raise NotFoundError("Test topilmadi.")
    if user.is_staff_wide or test.teacher_id == user.id:
        return test
    raise PermissionDeniedError("Bu test sizga tegishli emas.")


async def _assert_draft(test: Test) -> None:
    """Eʼlon qilingan testning savollari oʻzgartirilmaydi.

    Oʻquvchi ishlab boʻlgan savolni keyin tahrirlash natijani
    maʼnosiz qiladi: bir xil testni ikki xil odam ikki xil koʻrgan
    boʻlib chiqadi.
    """
    if test.status != TestStatus.DRAFT.value:
        raise ConflictError(
            "Eʼlon qilingan testning savollari oʻzgartirilmaydi. Avval qoralamaga qaytaring."
        )


# ─────────────────────────── Ustoz ───────────────────────────


async def _stats(
    session: AsyncSession, test_ids: list[uuid.UUID]
) -> dict[uuid.UUID, tuple[int, float | None]]:
    """Har test uchun (topshirganlar soni, oʻrtacha foiz) — bitta soʻrovda."""
    if not test_ids:
        return {}
    rows = (
        await session.execute(
            select(
                TestAttempt.test_id,
                func.count(func.distinct(TestAttempt.student_id)),
                func.avg(100.0 * TestAttempt.score / func.nullif(TestAttempt.max_score, 0)),
            )
            .where(
                TestAttempt.test_id.in_(test_ids),
                TestAttempt.submitted_at.is_not(None),
                TestAttempt.is_archived.is_(False),
            )
            .group_by(TestAttempt.test_id)
        )
    ).all()
    return {tid: (soni, round(float(o), 1) if o is not None else None) for tid, soni, o in rows}


async def _question_totals(
    session: AsyncSession, test_ids: list[uuid.UUID]
) -> dict[uuid.UUID, tuple[int, int]]:
    """(savollar soni, maksimal ball) — bitta soʻrovda."""
    if not test_ids:
        return {}
    rows = (
        await session.execute(
            select(
                TestQuestion.test_id,
                func.count(TestQuestion.id),
                func.coalesce(func.sum(TestQuestion.points), 0),
            )
            .where(TestQuestion.test_id.in_(test_ids), TestQuestion.is_archived.is_(False))
            .group_by(TestQuestion.test_id)
        )
    ).all()
    return {tid: (soni, int(ball)) for tid, soni, ball in rows}


async def _class_sizes(session: AsyncSession, class_ids: list[uuid.UUID]) -> dict[uuid.UUID, int]:
    if not class_ids:
        return {}
    rows = (
        await session.execute(
            select(Student.class_id, func.count(Student.id))
            .where(Student.class_id.in_(class_ids), Student.is_archived.is_(False))
            .group_by(Student.class_id)
        )
    ).all()
    return dict(rows)


def _row(
    test: Test,
    class_name: str,
    subject_name: str,
    totals: tuple[int, int],
    stats: tuple[int, float | None],
    class_size: int,
) -> TestRow:
    return TestRow(
        id=test.id,
        class_id=test.class_id,
        class_name=class_name,
        subject_id=test.subject_id,
        subject_name=subject_name,
        title=test.title,
        description=test.description or "",
        status=test.status,
        duration_minutes=test.duration_minutes,
        attempts_allowed=test.attempts_allowed,
        shuffle=test.shuffle,
        opens_at=test.opens_at,
        closes_at=test.closes_at,
        question_count=totals[0],
        max_score=totals[1],
        submitted_count=stats[0],
        total_students=class_size,
        average_percent=stats[1],
    )


async def teacher_tests(
    session: AsyncSession, user: CurrentUser, *, class_id: uuid.UUID | None = None
) -> list[TestRow]:
    """Ustozning testlari (TST-03)."""
    stmt = (
        select(Test, SchoolClass.name, Subject.name)
        .join(SchoolClass, SchoolClass.id == Test.class_id)
        .join(Subject, Subject.id == Test.subject_id)
        .where(Test.is_archived.is_(False))
        .order_by(Test.opens_at.desc())
    )
    if not user.is_staff_wide:
        stmt = stmt.where(Test.teacher_id == user.id)
    if class_id is not None:
        stmt = stmt.where(Test.class_id == class_id)

    rows = (await session.execute(stmt)).all()
    ids = [t.id for t, _, _ in rows]
    totals = await _question_totals(session, ids)
    stats = await _stats(session, ids)
    sizes = await _class_sizes(session, [t.class_id for t, _, _ in rows])

    return [
        _row(
            t,
            cls,
            subj,
            totals.get(t.id, (0, 0)),
            stats.get(t.id, (0, None)),
            sizes.get(t.class_id, 0),
        )
        for t, cls, subj in rows
    ]


async def create_test(
    session: AsyncSession,
    user: CurrentUser,
    *,
    class_id: uuid.UUID,
    subject_id: uuid.UUID,
    title: str,
    description: str = "",
    duration_minutes: int = 30,
    attempts_allowed: int = 1,
    shuffle: bool = True,
    opens_at: datetime,
    closes_at: datetime,
    ip: str | None = None,
) -> TestRow:
    """Yangi test — qoralama holatida (TST-03).

    Ustoz faqat oʻzi dars beradigan sinfda va oʻz fanidan test tuzadi —
    jurnal bilan bir xil qoida.
    """
    from app.services.grade_service import assert_teaches_class_subject

    await assert_teaches_class_subject(session, user, class_id, subject_id)

    if not title.strip():
        raise ValidationError("Test sarlavhasi boʻsh boʻlmasin.")
    if closes_at <= opens_at:
        raise ValidationError("Yopilish vaqti ochilishidan keyin boʻlsin.")
    if duration_minutes < 1:
        raise ValidationError("Davomiylik 1 daqiqadan kam boʻlmasin.")
    if attempts_allowed < 1:
        raise ValidationError("Urinishlar soni 1 dan kam boʻlmasin.")

    cls = await session.get(SchoolClass, class_id)
    subject = await session.get(Subject, subject_id)
    if cls is None or cls.is_archived:
        raise NotFoundError("Sinf topilmadi.")
    if subject is None or subject.is_archived:
        raise NotFoundError("Fan topilmadi.")

    test = Test(
        class_id=class_id,
        subject_id=subject_id,
        teacher_id=user.id,
        title=title.strip(),
        description=(description or "").strip(),
        duration_minutes=duration_minutes,
        attempts_allowed=attempts_allowed,
        shuffle=shuffle,
        opens_at=opens_at,
        closes_at=closes_at,
    )
    session.add(test)
    await session.flush()

    audit_service.record(
        session,
        object_type="test",
        object_id=test.id,
        action=AuditAction.CREATE,
        new={"title": test.title, "class_id": class_id},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()

    sizes = await _class_sizes(session, [class_id])
    return _row(test, cls.name, subject.name, (0, 0), (0, None), sizes.get(class_id, 0))


async def set_status(
    session: AsyncSession,
    user: CurrentUser,
    test_id: uuid.UUID,
    status: str,
    *,
    ip: str | None = None,
) -> TestRow:
    """Qoralama → eʼlon → yakunlangan (TST-03).

    Savolsiz testni eʼlon qilib boʻlmaydi: oʻquvchi boʻsh ekran koʻrardi.
    """
    if status not in {s.value for s in TestStatus}:
        raise ValidationError("Nomaʼlum holat.")

    test = await _load_for_teacher(session, user, test_id)
    totals = await _question_totals(session, [test.id])
    soni = totals.get(test.id, (0, 0))[0]

    if status == TestStatus.PUBLISHED.value and soni == 0:
        raise ValidationError("Savolsiz testni eʼlon qilib boʻlmaydi.")

    eski = test.status
    if eski == status:
        return await _one(session, test)

    test.status = status
    audit_service.record(
        session,
        object_type="test",
        object_id=test.id,
        action=AuditAction.UPDATE,
        old={"status": eski},
        new={"status": status},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()
    return await _one(session, test)


async def _one(session: AsyncSession, test: Test) -> TestRow:
    cls = await session.get(SchoolClass, test.class_id)
    subject = await session.get(Subject, test.subject_id)
    totals = await _question_totals(session, [test.id])
    stats = await _stats(session, [test.id])
    sizes = await _class_sizes(session, [test.class_id])
    return _row(
        test,
        cls.name if cls else "",
        subject.name if subject else "",
        totals.get(test.id, (0, 0)),
        stats.get(test.id, (0, None)),
        sizes.get(test.class_id, 0),
    )


async def archive_test(
    session: AsyncSession, user: CurrentUser, test_id: uuid.UUID, *, ip: str | None = None
) -> None:
    """Oʻchirish YOʻQ (CLAUDE.md 1-qoida) — oʻtgan urinishlar qoladi."""
    test = await _load_for_teacher(session, user, test_id)
    test.is_archived = True
    test.archived_at = utcnow()
    audit_service.record(
        session,
        object_type="test",
        object_id=test.id,
        action=AuditAction.ARCHIVE,
        old={"title": test.title},
        new={"is_archived": True},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()


# ─────────────────────────── Savollar ───────────────────────────


@dataclass(frozen=True, slots=True)
class OptionInput:
    text: str
    is_correct: bool


async def teacher_questions(
    session: AsyncSession, user: CurrentUser, test_id: uuid.UUID
) -> list[QuestionRow]:
    """Ustoz koʻrinishi — toʻgʻri javoblar BILAN."""
    await _load_for_teacher(session, user, test_id)

    rows = list(
        (
            await session.execute(
                select(TestQuestion)
                .options(selectinload(TestQuestion.options))
                .where(TestQuestion.test_id == test_id, TestQuestion.is_archived.is_(False))
                .order_by(TestQuestion.position)
            )
        ).scalars()
    )
    return [
        QuestionRow(
            id=q.id,
            position=q.position,
            text=q.text,
            kind=q.kind,
            points=q.points,
            options=[
                OptionRow(id=o.id, text=o.text, is_correct=o.is_correct)
                for o in q.options
                if not o.is_archived
            ],
        )
        for q in rows
    ]


def _validate_options(kind: str, options: list[OptionInput]) -> None:
    if len(options) < 2:
        raise ValidationError("Savolda kamida ikkita variant boʻlsin.")
    if len(options) > MAX_OPTIONS:
        raise ValidationError(f"Variantlar soni {MAX_OPTIONS} tadan oshmasin.")
    if any(not o.text.strip() for o in options):
        raise ValidationError("Variant matni boʻsh boʻlmasin.")

    togri = sum(1 for o in options if o.is_correct)
    if togri == 0:
        raise ValidationError("Kamida bitta toʻgʻri javob belgilansin.")
    if kind == QuestionKind.SINGLE.value and togri > 1:
        raise ValidationError("«Bitta javob» turida faqat bitta variant toʻgʻri boʻlishi mumkin.")
    if togri == len(options):
        raise ValidationError("Hamma variant toʻgʻri boʻlsa savolning maʼnosi qolmaydi.")


async def add_question(
    session: AsyncSession,
    user: CurrentUser,
    test_id: uuid.UUID,
    *,
    text: str,
    kind: str,
    points: int,
    options: list[OptionInput],
    ip: str | None = None,
) -> QuestionRow:
    """TST-01, TST-02: savol qoʻshadi."""
    test = await _load_for_teacher(session, user, test_id)
    await _assert_draft(test)

    if not text.strip():
        raise ValidationError("Savol matni boʻsh boʻlmasin.")
    if kind not in {k.value for k in QuestionKind}:
        raise ValidationError("Nomaʼlum savol turi.")
    if points < 1:
        raise ValidationError("Ball 1 dan kichik boʻlmasin.")
    _validate_options(kind, options)

    soni = await session.scalar(
        select(func.count(TestQuestion.id)).where(
            TestQuestion.test_id == test_id, TestQuestion.is_archived.is_(False)
        )
    )
    if (soni or 0) >= MAX_QUESTIONS:
        raise ValidationError(f"Savollar soni {MAX_QUESTIONS} tadan oshmasin.")

    question = TestQuestion(
        test_id=test_id,
        position=(soni or 0) + 1,
        text=text.strip(),
        kind=kind,
        points=points,
    )
    session.add(question)
    await session.flush()

    for i, o in enumerate(options, start=1):
        session.add(
            TestOption(
                question_id=question.id,
                position=i,
                text=o.text.strip(),
                is_correct=o.is_correct,
            )
        )

    audit_service.record(
        session,
        object_type="test_question",
        object_id=question.id,
        action=AuditAction.CREATE,
        new={"test_id": test_id, "points": points},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()

    rows = await teacher_questions(session, user, test_id)
    return next(q for q in rows if q.id == question.id)


async def archive_question(
    session: AsyncSession,
    user: CurrentUser,
    question_id: uuid.UUID,
    *,
    ip: str | None = None,
) -> None:
    """Savolni roʻyxatdan chiqaradi. Oʻchirish YOʻQ (1-qoida)."""
    question = await session.get(TestQuestion, question_id)
    if question is None:
        raise NotFoundError("Savol topilmadi.")

    test = await _load_for_teacher(session, user, question.test_id)
    await _assert_draft(test)

    if question.is_archived:
        return
    question.is_archived = True
    question.archived_at = utcnow()

    audit_service.record(
        session,
        object_type="test_question",
        object_id=question.id,
        action=AuditAction.ARCHIVE,
        old={"text": question.text[:100]},
        new={"is_archived": True},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()


# ─────────────────────── Oʻquvchi koʻrinishi ───────────────────────


async def available_tests(
    session: AsyncSession, user: CurrentUser, student_id: uuid.UUID
) -> list[TestRow]:
    """Oʻquvchiga ochiq testlar (TST-04).

    Faqat eʼlon qilingan va oʻz sinfiniki. Kesim soʻrov darajasida —
    ota-ona URL dagi `student_id` ni oʻzgartirsa `403` (X-1).
    """
    await assert_can_view_student(session, user, student_id)

    student = await session.get(Student, student_id)
    if student is None or student.class_id is None:
        return []

    rows = (
        await session.execute(
            select(Test, SchoolClass.name, Subject.name)
            .join(SchoolClass, SchoolClass.id == Test.class_id)
            .join(Subject, Subject.id == Test.subject_id)
            .where(
                Test.class_id == student.class_id,
                Test.is_archived.is_(False),
                Test.status == TestStatus.PUBLISHED.value,
            )
            .order_by(Test.closes_at)
        )
    ).all()

    ids = [t.id for t, _, _ in rows]
    totals = await _question_totals(session, ids)
    sizes = await _class_sizes(session, [student.class_id])

    return [
        _row(t, cls, subj, totals.get(t.id, (0, 0)), (0, None), sizes.get(student.class_id, 0))
        for t, cls, subj in rows
    ]


async def _student_questions(
    session: AsyncSession, test: Test, *, shuffle: bool
) -> list[QuestionRow]:
    """Oʻquvchi koʻradigan savollar — TOʻGʻRI JAVOBSIZ (1-qoida).

    `is_correct` soʻrovga umuman qoʻshilmaydi: sxemada kesib tashlashga
    tayanilmaydi, ustun ORM obyektiga ham tushmaydi.
    """
    savollar = list(
        (
            await session.execute(
                select(TestQuestion)
                .where(TestQuestion.test_id == test.id, TestQuestion.is_archived.is_(False))
                .order_by(TestQuestion.position)
            )
        ).scalars()
    )
    if not savollar:
        return []

    variantlar = (
        await session.execute(
            select(TestOption.id, TestOption.question_id, TestOption.text, TestOption.position)
            .where(
                TestOption.question_id.in_([q.id for q in savollar]),
                TestOption.is_archived.is_(False),
            )
            .order_by(TestOption.position)
        )
    ).all()

    by_question: dict[uuid.UUID, list[OptionRow]] = {}
    for oid, qid, matn, _ in variantlar:
        by_question.setdefault(qid, []).append(OptionRow(id=oid, text=matn, is_correct=None))

    natija = [
        QuestionRow(
            id=q.id,
            position=q.position,
            text=q.text,
            kind=q.kind,
            points=q.points,
            options=by_question.get(q.id, []),
        )
        for q in savollar
    ]
    if shuffle:
        random.shuffle(natija)
    return natija


async def start_attempt(
    session: AsyncSession,
    user: CurrentUser,
    test_id: uuid.UUID,
    student_id: uuid.UUID,
    *,
    ip: str | None = None,
) -> tuple[TestAttempt, list[QuestionRow]]:
    """Urinishni boshlaydi va savollarni qaytaradi (TST-04).

    Urinishlar soni SERVERDA cheklanadi (3-qoida).
    """
    # Faqat oʻquvchining oʻzi boshlaydi — vasiy/ustoz uning nomidan emas.
    await assert_student_self(session, user, student_id)

    test = await session.get(Test, test_id)
    if test is None or test.is_archived:
        raise NotFoundError("Test topilmadi.")
    if test.status != TestStatus.PUBLISHED.value:
        raise ConflictError("Test hozir faol emas.")

    student = await session.get(Student, student_id)
    if student is None or student.class_id != test.class_id:
        raise PermissionDeniedError("Bu test sizning sinfingizga tegishli emas.")

    hozir = utcnow()
    if hozir < test.opens_at:
        raise ConflictError("Test hali ochilmagan.")
    if hozir > test.closes_at:
        raise ConflictError("Test yopilgan.")

    oldingi = list(
        (
            await session.execute(
                select(TestAttempt)
                .where(
                    TestAttempt.test_id == test_id,
                    TestAttempt.student_id == student_id,
                    TestAttempt.is_archived.is_(False),
                )
                .order_by(TestAttempt.attempt_no)
            )
        ).scalars()
    )

    # Tugallanmagan urinish bor boʻlsa — oʻshanisi davom etadi. Sahifa
    # yangilanganda yangi urinish sarflanib ketmasin.
    ochiq = next((a for a in oldingi if a.submitted_at is None), None)
    if ochiq is not None:
        return ochiq, await _student_questions(session, test, shuffle=test.shuffle)

    if len(oldingi) >= test.attempts_allowed:
        raise ConflictError(
            f"Urinishlar tugadi ({test.attempts_allowed} tadan {len(oldingi)} ta ishlatilgan)."
        )

    totals = await _question_totals(session, [test.id])
    attempt = TestAttempt(
        test_id=test_id,
        student_id=student_id,
        attempt_no=len(oldingi) + 1,
        started_at=hozir,
        max_score=totals.get(test.id, (0, 0))[1],
    )
    session.add(attempt)
    await session.flush()

    audit_service.record(
        session,
        object_type="test_attempt",
        object_id=attempt.id,
        action=AuditAction.CREATE,
        new={"test_id": test_id, "student_id": student_id, "attempt_no": attempt.attempt_no},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()

    return attempt, await _student_questions(session, test, shuffle=test.shuffle)


@dataclass(frozen=True, slots=True)
class AnswerInput:
    question_id: uuid.UUID
    selected: list[uuid.UUID]


async def submit_attempt(
    session: AsyncSession,
    user: CurrentUser,
    attempt_id: uuid.UUID,
    answers: list[AnswerInput],
    *,
    ip: str | None = None,
) -> AttemptRow:
    """Javoblarni qabul qiladi va AVTOMATIK tekshiradi (TST-04).

    Ball frontenddan olinmaydi — server toʻgʻri javoblar bilan
    solishtirib oʻzi hisoblaydi (2-qoida).
    """
    attempt = await session.get(TestAttempt, attempt_id)
    if attempt is None or attempt.is_archived:
        raise NotFoundError("Urinish topilmadi.")

    # Faqat oʻquvchining oʻzi yakunlaydi (K6 — natija butunligi).
    await assert_student_self(session, user, attempt.student_id)

    if attempt.submitted_at is not None:
        raise ConflictError("Bu urinish allaqachon yakunlangan.")

    test = await session.get(Test, attempt.test_id)
    if test is None:
        raise NotFoundError("Test topilmadi.")

    # Test yopilgandan keyin ham javob qabul qilinadi: oʻquvchi
    # ishlab oʻtirganda muddat tugab qolishi mumkin va uning mehnati
    # yoʻqolmasligi kerak. Yangi urinish esa boshlanmaydi
    # (`start_attempt` ni koʻring).
    hozir = utcnow()

    savollar = {
        q.id: q
        for q in (
            await session.execute(
                select(TestQuestion).where(
                    TestQuestion.test_id == test.id, TestQuestion.is_archived.is_(False)
                )
            )
        ).scalars()
    }

    togri_javoblar: dict[uuid.UUID, set[uuid.UUID]] = {}
    for oid, qid, is_correct in (
        await session.execute(
            select(TestOption.id, TestOption.question_id, TestOption.is_correct).where(
                TestOption.question_id.in_(savollar.keys()),
                TestOption.is_archived.is_(False),
            )
        )
    ).all():
        if is_correct:
            togri_javoblar.setdefault(qid, set()).add(oid)

    berilgan = {a.question_id: set(a.selected) for a in answers}

    ball = 0
    for qid, q in savollar.items():
        tanlangan = berilgan.get(qid, set())
        kerak = togri_javoblar.get(qid, set())
        # Toʻliq moslik: yarim javobga ball berilmaydi. Bu eng oddiy va
        # eng tushunarli qoida — oʻquvchi nega 0,5 ball olganini
        # soʻramaydi.
        mos = bool(kerak) and tanlangan == kerak
        session.add(
            TestAnswer(
                attempt_id=attempt.id,
                question_id=qid,
                selected=sorted(tanlangan),
                is_correct=mos,
                points_awarded=q.points if mos else 0,
            )
        )
        if mos:
            ball += q.points

    attempt.score = ball
    attempt.max_score = sum(q.points for q in savollar.values())
    attempt.submitted_at = hozir

    audit_service.record(
        session,
        object_type="test_attempt",
        object_id=attempt.id,
        action=AuditAction.UPDATE,
        new={"score": ball, "max_score": attempt.max_score},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()

    student = await session.get(Student, attempt.student_id)
    return _attempt_row(attempt, student.full_name if student else "")


def _attempt_row(a: TestAttempt, full_name: str) -> AttemptRow:
    foiz = (
        round(100.0 * a.score / a.max_score, 1) if a.score is not None and a.max_score > 0 else None
    )
    return AttemptRow(
        id=a.id,
        test_id=a.test_id,
        student_id=a.student_id,
        full_name=full_name,
        attempt_no=a.attempt_no,
        started_at=a.started_at,
        submitted_at=a.submitted_at,
        score=a.score,
        max_score=a.max_score,
        percent=foiz,
    )


async def test_results(
    session: AsyncSession, user: CurrentUser, test_id: uuid.UUID
) -> list[AttemptRow]:
    """TST-05: natijalar — ustoz uchun."""
    await _load_for_teacher(session, user, test_id)

    rows = (
        await session.execute(
            select(TestAttempt, Student)
            .join(Student, Student.id == TestAttempt.student_id)
            .where(TestAttempt.test_id == test_id, TestAttempt.is_archived.is_(False))
            .order_by(Student.last_name, TestAttempt.attempt_no)
        )
    ).all()
    return [_attempt_row(a, s.full_name) for a, s in rows]


async def student_attempts(
    session: AsyncSession, user: CurrentUser, student_id: uuid.UUID
) -> list[AttemptRow]:
    """Oʻquvchining oʻz natijalari (TST-05).

    Ota-ona faqat oʻz farzandiniki — tekshiruv `access.py` da (X-1).
    """
    await assert_can_view_student(session, user, student_id)

    rows = (
        await session.execute(
            select(TestAttempt, Student)
            .join(Student, Student.id == TestAttempt.student_id)
            .where(
                TestAttempt.student_id == student_id,
                TestAttempt.is_archived.is_(False),
                TestAttempt.submitted_at.is_not(None),
            )
            .order_by(TestAttempt.submitted_at.desc())
        )
    ).all()
    return [_attempt_row(a, s.full_name) for a, s in rows]


# ─────────────── Savollarni Excel'dan import (TST-06) ───────────────

#: Shablon ustunlari. Tartib MUHIM — import shu tartibda oʻqiydi.
QUESTION_COLUMNS = [
    ("Savol", "Savol matni (majburiy)"),
    ("Ball", "Butun son, sukut boʻyicha 1"),
    ("Variant 1", "Toʻgʻri variantni «+» bilan boshlang: «+ Toshkent»"),
    ("Variant 2", "Kamida ikkita variant boʻlsin"),
    ("Variant 3", "Boʻsh qoldirilsa oʻtkazib yuboriladi"),
    ("Variant 4", ""),
    ("Variant 5", ""),
    ("Variant 6", ""),
]

#: Toʻgʻri javob belgisi. `+` tanlangan, chunki u Excel katakchasida
#: koʻzga tashlanadi va formulaga aylanib ketmaydi (`=` dan farqli).
CORRECT_MARK = "+"


def build_question_template() -> bytes:
    """Savollar uchun boʻsh Excel shablon (TST-06).

    Savol turi ustuni ATAYLAB yoʻq: u toʻgʻri javoblar sonidan
    kelib chiqadi — bitta boʻlsa «single», bir nechta boʻlsa
    «multiple». Foydalanuvchi bir joyda «multiple» deb yozib,
    boshqa joyda bitta javob belgilab qoʻyishi eng koʻp uchraydigan
    xato edi.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Savollar"
    ws.append([c[0] for c in QUESTION_COLUMNS])
    ws.append([
        "Oʻzbekiston poytaxti qaysi shahar?",
        1,
        f"{CORRECT_MARK} Toshkent",
        "Samarqand",
        "Buxoro",
        "Xiva",
        "",
        "",
    ])
    ws.append([
        "Qaysilari suyuqlik? (bir nechta javob)",
        2,
        f"{CORRECT_MARK} Suv",
        f"{CORRECT_MARK} Sut",
        "Temir",
        "Yogʻoch",
        "",
        "",
    ])
    for i, _ in enumerate(QUESTION_COLUMNS, start=1):
        ws.column_dimensions[chr(64 + i)].width = 34

    y = wb.create_sheet("Yoʻriqnoma")
    y.column_dimensions["A"].width = 90
    y.append(["Savollar shabloni — toʻldirish qoidalari"])
    y.append([""])
    for nom, izoh in QUESTION_COLUMNS:
        if izoh:
            y.append([f"• {nom}: {izoh}"])
    y.append([""])
    y.append([f"Toʻgʻri variant «{CORRECT_MARK}» belgisi bilan boshlanadi."])
    y.append(["Savol turi avtomatik: bitta toʻgʻri javob — «Bitta javob»,"])
    y.append(["bir nechtasi — «Bir nechta javob»."])
    y.append(["Hamma variant toʻgʻri boʻlsa savol qabul qilinmaydi."])
    y.append(["Import faqat QORALAMA testga qilinadi."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass(frozen=True, slots=True)
class QuestionImportResult:
    added: int
    warnings: list[str]


def _parse_questions_xlsx(data: bytes) -> tuple[list[dict], list[str]]:
    """Shablonni oʻqiydi. Buzuq qatorni TASHLAB ketadi, xato bermaydi.

    Sabab: 60 ta savolli fayldagi bitta buzuq qator butun importni
    toʻxtatsa, foydalanuvchi qaysi qator ekanini topolmay qiynaladi.
    Har tashlangan qator ogohlantirishda nomi bilan qaytadi.
    """
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — foydalanuvchi fayli, har xil buzilishi mumkin
        raise ValidationError("Fayl ochilmadi — .xlsx shablon yuklang.") from e

    ws = wb["Savollar"] if "Savollar" in wb.sheetnames else wb.active
    warnings: list[str] = []
    savollar: list[dict] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        qiymatlar = (list(row) + [None] * len(QUESTION_COLUMNS))[: len(QUESTION_COLUMNS)]
        matn, ball_raw = qiymatlar[0], qiymatlar[1]
        xom_variantlar = qiymatlar[2:]

        if matn is None or not str(matn).strip():
            warnings.append(f"{idx}-qator: savol matni boʻsh — oʻtkazib yuborildi.")
            continue

        try:
            ball = int(ball_raw) if ball_raw is not None else 1
        except (TypeError, ValueError):
            warnings.append(f"{idx}-qator: ball son emas — 1 deb olindi.")
            ball = 1
        if ball < 1:
            warnings.append(f"{idx}-qator: ball 1 dan kichik — 1 deb olindi.")
            ball = 1

        variantlar: list[OptionInput] = []
        for xom in xom_variantlar:
            if xom is None or not str(xom).strip():
                continue
            matn_v = str(xom).strip()
            togri = matn_v.startswith(CORRECT_MARK)
            if togri:
                matn_v = matn_v[len(CORRECT_MARK) :].strip()
            if not matn_v:
                continue
            variantlar.append(OptionInput(text=matn_v, is_correct=togri))

        togri_soni = sum(1 for v in variantlar if v.is_correct)
        # Tur toʻgʻri javoblar sonidan kelib chiqadi — foydalanuvchi
        # uni alohida yozmaydi va shu bilan ziddiyat ham chiqmaydi.
        kind = (
            QuestionKind.MULTIPLE.value if togri_soni > 1 else QuestionKind.SINGLE.value
        )

        try:
            _validate_options(kind, variantlar)
        except ValidationError as e:
            warnings.append(f"{idx}-qator: {e.message}")
            continue

        savollar.append(
            {
                "text": str(matn).strip(),
                "kind": kind,
                "points": ball,
                "options": variantlar,
            }
        )

    if not savollar:
        raise ValidationError("Faylda birorta ham yaroqli savol topilmadi.")
    return savollar, warnings


async def import_questions(
    session: AsyncSession,
    user: CurrentUser,
    test_id: uuid.UUID,
    *,
    data: bytes,
    ip: str | None = None,
) -> QuestionImportResult:
    """TST-06: savollarni Excel shablonidan ommaviy import qiladi.

    Import faqat QORALAMA testga — `add_question` bilan bir xil qoida.
    Savollar mavjudlariga QOʻSHILADI, ularni almashtirmaydi: fayl ikki
    marta yuklansa savollar ikkilanadi, lekin «hammasini oʻchirib
    qayta yozish» xatosi qaytarib boʻlmaydigan yoʻqotish boʻlardi.
    """
    test = await _load_for_teacher(session, user, test_id)
    await _assert_draft(test)

    savollar, warnings = _parse_questions_xlsx(data)

    mavjud = await session.scalar(
        select(func.count(TestQuestion.id)).where(
            TestQuestion.test_id == test_id, TestQuestion.is_archived.is_(False)
        )
    )
    joriy = mavjud or 0
    if joriy + len(savollar) > MAX_QUESTIONS:
        raise ValidationError(
            f"Testda {MAX_QUESTIONS} tadan koʻp savol boʻlmaydi. "
            f"Hozir {joriy} ta bor, faylda {len(savollar)} ta."
        )

    for n, s in enumerate(savollar, start=1):
        question = TestQuestion(
            test_id=test_id,
            position=joriy + n,
            text=s["text"],
            kind=s["kind"],
            points=s["points"],
        )
        session.add(question)
        await session.flush()
        for i, o in enumerate(s["options"], start=1):
            session.add(
                TestOption(
                    question_id=question.id,
                    position=i,
                    text=o.text,
                    is_correct=o.is_correct,
                )
            )

    audit_service.record(
        session,
        object_type="test",
        object_id=test_id,
        action=AuditAction.UPDATE,
        new={"imported_questions": len(savollar)},
        actor_id=user.id,
        ip=ip,
    )
    await session.commit()
    return QuestionImportResult(added=len(savollar), warnings=warnings)
