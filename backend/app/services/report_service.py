"""Hisobotlarni Excel'ga eksport (DIR-08, X-13).

Uchta hisobot: sinflar kesimi (DIR-03), ustozlar faoliyati (DIR-04) va
qarzdorlik (DIR-06). PDF alohida yozilmagan — brauzerning oʻz chop
etish oynasi ishlatiladi (`print-doc`, `globals.css`), chunki
hisobotning ekrandagi koʻrinishi allaqachon tayyor va uni ikkinchi
marta PDF kutubxonasida qayta chizishning maʼnosi yoʻq.

**X-13: har eksport auditga tushadi.** Eng ehtimolli sizib chiqish —
hujum emas, xodim. «Kim, qachon, qaysi roʻyxatni yuklab oldi» savoli
javobsiz qolmasin. Shuning uchun audit yozuvi eksportdan OLDIN emas,
fayl tayyor boʻlgach yoziladi va u bilan bitta tranzaksiyada
saqlanadi.

**Huquq: `reports.export`.** Rol yetarli emas — direktor hisobotni
koʻradi, lekin yuklab olish alohida beriladigan huquq.
"""

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ValidationError
from app.core.timeutil import to_display, utcnow
from app.models import AuditAction, Permission
from app.services import audit_service, director_service, payment_service
from app.services.access import CurrentUser
from app.services.permissions import assert_permission

#: Qaysi hisobotlar eksport qilinadi. Kalit URL da keladi.
REPORT_KINDS = ("sinflar", "ustozlar", "qarzdorlik")

REPORT_TITLES: dict[str, str] = {
    "sinflar": "Sinflar kesimi",
    "ustozlar": "Ustozlar faoliyati",
    "qarzdorlik": "Toʻlov va qarzdorlik",
}


@dataclass(frozen=True, slots=True)
class Report:
    filename: str
    content: bytes


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    """Bitta varaq: sarlavha qatori qalin va muzlatilgan.

    Muzlatish shunchaki qulaylik emas: 100 qatorli qarzdorlik
    roʻyxatida ustun nomlari koʻrinmasa, «-2 300 000» qaysi ustun
    ekani anglashilmaydi.
    """
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = max(14, min(40, len(h) + 8))


def _meta_sheet(wb: Workbook, kind: str, actor: CurrentUser, row_count: int) -> None:
    """«Qachon, kim, nechta qator» varagʻi.

    Fayl qoʻldan qoʻlga oʻtadi va bir oydan keyin «bu qaysi sanadagi
    holat?» degan savol chiqadi. Sana faylning ichida turishi kerak.
    """
    ws = wb.active
    ws.title = "Maʼlumot"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 44
    for nom, qiymat in [
        ("Hisobot", REPORT_TITLES[kind]),
        ("Yuklab olindi", to_display(utcnow()).strftime("%Y-%m-%d %H:%M")),
        ("Vaqt mintaqasi", "Asia/Tashkent"),
        ("Kim yuklab oldi", actor.full_name),
        ("Qatorlar soni", row_count),
    ]:
        ws.append([nom, qiymat])
    ws["A1"].font = Font(bold=True)


async def _rows_for(
    session: AsyncSession, actor: CurrentUser, kind: str
) -> tuple[list[str], list[list]]:
    """Hisobot qatorlari. Har biri OʻZ servisidan oladi.

    Sanoq va foiz bu yerda QAYTA hisoblanmaydi — ekranda koʻringan son
    bilan fayldagi son bir xil boʻlishi shart, aks holda ular
    solishtirilganda ishonch yoʻqoladi.
    """
    if kind == "sinflar":
        rows = await director_service.classes(session)
        return (
            ["Sinf", "Sinf rahbari", "Oʻquvchi", "Davomat %", "Davomat yozuvi", "Oʻrtacha baho"],
            [
                [
                    r.name,
                    r.homeroom_teacher_name or "—",
                    r.student_count,
                    r.attendance_percent,
                    r.attendance_records,
                    r.average_grade,
                ]
                for r in rows
            ],
        )

    if kind == "ustozlar":
        rows = await director_service.teachers(session)
        return (
            [
                "Ustoz", "Fanlar", "Sinf rahbari", "Haftalik soat",
                "Rejadagi dars", "Davomat belgilangan", "Qoʻyilgan baho",
                "Oʻrtacha baho", "Imtihon", "Uy vazifasi",
            ],
            [
                [
                    r.full_name,
                    ", ".join(r.subjects) or "—",
                    r.homeroom_class_name or "—",
                    r.weekly_hours,
                    r.lessons_planned,
                    r.lessons_with_attendance,
                    r.grades_given,
                    r.average_grade_given,
                    r.exams_held,
                    r.homework_given,
                ]
                for r in rows
            ],
        )

    # qarzdorlik
    rows = await payment_service.finance_rows(session, actor)
    return (
        ["Oʻquvchi", "Sinf", "Oylik", "Hisoblangan", "Toʻlangan", "Balans", "Holat", "Ketgan"],
        [
            [
                r.student_name,
                r.class_name or "—",
                r.monthly_fee or 0,
                r.charged,
                r.paid,
                r.balance,
                r.status,
                "ha" if r.is_archived else "",
            ]
            for r in rows
        ],
    )


async def export_xlsx(
    session: AsyncSession,
    actor: CurrentUser,
    *,
    kind: str,
    ip: str | None = None,
) -> Report:
    """Hisobotni Excel qilib beradi va eksportni auditga yozadi (DIR-08, X-13)."""
    if kind not in REPORT_KINDS:
        raise ValidationError("Nomaʼlum hisobot turi.")

    await assert_permission(session, actor, Permission.REPORTS_EXPORT)

    headers, rows = await _rows_for(session, actor, kind)

    wb = Workbook()
    _meta_sheet(wb, kind, actor, len(rows))
    _sheet(wb, REPORT_TITLES[kind], headers, rows)
    buf = io.BytesIO()
    wb.save(buf)

    audit_service.record(
        session,
        object_type="report_export",
        action=AuditAction.CREATE,
        new={"kind": kind, "rows": len(rows), "format": "xlsx"},
        actor_id=actor.id,
        ip=ip,
    )
    await session.commit()

    sana = to_display(utcnow()).strftime("%Y-%m-%d")
    return Report(
        filename=f"tarbion-{kind}-{sana}.xlsx",
        content=buf.getvalue(),
    )
