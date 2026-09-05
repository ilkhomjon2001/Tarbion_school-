"""Oʻquv rejalari (metodik baza) — CRUD + Excel import/eksport.

Oqim: shablonni yuklab olish → toʻldirish → import (QORALAMA) →
koʻrib chiqish → JORIY qilish (ustozlarga koʻrinadi). Bir (fan, yil,
sinf) uchun bitta joriy reja — yangisi joriy boʻlganda eskisi ARXIV.

Kirish: yozish amallari oʻquv boʻlimi/admin/superadmin (router darajasida
rol bilan), joriy rejalarni OʻQISH esa barcha xodimlarga ochiq (ustoz
kabineti shu yerdan oladi).
"""

import io
import re
import uuid
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import (
    ConflictError,
    NotFoundError,
    PermissionDeniedError,
    ValidationError,
)
from app.core.timeutil import utcnow
from app.models import (
    PROGRAM_YEARS,
    AuditAction,
    CurriculumPlan,
    CurriculumStatus,
    RoleName,
    SchoolSettings,
)
from app.services import audit_service, storage
from app.services.access import CurrentUser

#: Shablon ustunlari — tartib MUHIM (import shu tartibda oʻqiydi).
COLUMNS = [
    ("Chorak", "1–4 raqami"),
    ("Mavzu", "Dars mavzusi (majburiy)"),
    ("Tur", "qurish / dasturlash / nazorat / loyiha / elektronika / ai"),
    ("Model", "Model nomi (ixtiyoriy)"),
    ("Maqsad", "Har qatorda bitta maqsad (Alt+Enter bilan)"),
    ("Lugʻat", "Har qatorda bitta atama"),
    ("Nazariya", "Nazariy qism — har qatorda bitta band"),
    ("Amaliy", "Amaliy qism — har qatorda bitta band"),
    ("Uyga vazifa", "Har qatorda bitta topshiriq"),
    ("Kutilayotgan natija", "Dars oxirida oʻquvchi nima qila oladi (MET-02)"),
    ("Kerakli jihozlar", "Har qatorda bitta jihoz — qidiruvga tushadi (MET-05)"),
    ("Baholash mezoni", "Nima boʻyicha baholanadi — har qatorda bitta mezon"),
    ("Resurslar", "Har qatorda bitta resurs"),
    ("Video havola", "YouTube va shu kabi — kartochka ichida koʻrsatiladi (MET-04)"),
]

TYPES = {
    "qurish", "dasturlash", "nazorat", "loyiha",
    "elektronika", "ai", "spike", "arduino", "esp32",
}

MAX_LESSONS = 200

_APO = re.compile(r"([oOgG])['’`ʼ]")
_REST = re.compile(r"['’`]")


def _toza(s: str) -> str:
    """Apostrof normalizatsiyasi (CLAUDE.md 8-qoida)."""
    s = _APO.sub(lambda m: m.group(1) + "ʻ", s.strip())
    return _REST.sub("ʼ", s)


def _lines(cell: object) -> list[str]:
    if cell is None:
        return []
    return [_toza(x) for x in str(cell).splitlines() if x.strip()]


@dataclass
class ImportResult:
    plan: CurriculumPlan
    warnings: list[str] = field(default_factory=list)


# ─────────────────────────── Shablon ───────────────────────────


def build_template() -> bytes:
    """Boʻsh Excel shablon — «Reja» varagʻi + «Yoʻriqnoma»."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reja"
    ws.append([c[0] for c in COLUMNS])
    # Namuna qator — foydalanuvchi formatni koʻrsin.
    ws.append([
        1,
        "Kirish darsi: fan bilan tanishuv",
        "qurish",
        "",
        "Oʻquvchilar fan nimani oʻrganishini biladilar",
        "Atama (Term) — izohi",
        "Kirish suhbati (10 daqiqa)\nAsosiy tushunchalar (20 daqiqa)",
        "Amaliy mashq (15 daqiqa)",
        "Oʻtilganlarni takrorlash",
        "Oʻquvchi fanning uchta asosiy tushunchasini ayta oladi",
        "Doska\nProyektor\nTarqatma material",
        "Savollarga javob berish faolligi\nAmaliy mashq natijasi",
        "Darslik, 5–12-betlar",
        "https://www.youtube.com/watch?v=...",
    ])
    for i, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[chr(64 + i)].width = 28

    y = wb.create_sheet("Yoʻriqnoma")
    y.column_dimensions["A"].width = 90
    y.append(["Oʻquv rejasi shabloni — toʻldirish qoidalari"])
    y.append([""])
    for nom, izoh in COLUMNS:
        y.append([f"• {nom}: {izoh}"])
    y.append([""])
    y.append(["Bir katakda bir nechta band boʻlsa — har birini yangi qatordan (Alt+Enter)."])
    y.append(["Chorak 1 dan 4 gacha; qatorlar tartibi darslar tartibini belgilaydi."])
    y.append([
        "Fayl yuklangach reja QORALAMA boʻladi — "
        "«Joriy qilish»dan keyin ustozlarga koʻrinadi."
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────── Import ───────────────────────────


def _parse_xlsx(data: bytes) -> tuple[list[dict], list[str]]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — foydalanuvchi faylı, har xil buzilishi mumkin
        raise ValidationError("Fayl ochilmadi — .xlsx shablon yuklang.") from e

    ws = wb["Reja"] if "Reja" in wb.sheetnames else wb.active
    warnings: list[str] = []
    lessons: list[dict] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        (
            chorak_raw,
            mavzu,
            tur,
            model,
            maqsad,
            lugat,
            nazariya,
            amaliy,
            uyga,
            natija,
            jihoz,
            baholash,
            resurslar,
            video,
        ) = (list(row) + [None] * len(COLUMNS))[: len(COLUMNS)]

        if mavzu is None or not str(mavzu).strip():
            warnings.append(f"{idx}-qator: mavzu boʻsh — oʻtkazib yuborildi.")
            continue
        try:
            chorak = int(chorak_raw)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            chorak = 0
        if chorak < 1 or chorak > 4:
            warnings.append(f"{idx}-qator: chorak 1–4 boʻlishi kerak — 1 deb olindi.")
            chorak = 1

        tur_s = str(tur).strip().lower() if tur else "qurish"
        if tur_s not in TYPES:
            warnings.append(f"{idx}-qator: nomaʼlum tur «{tur_s}» — «qurish» deb olindi.")
            tur_s = "qurish"

        dars: dict = {
            "chorak": chorak,
            "title": _toza(str(mavzu)),
            "type": tur_s,
            "model": _toza(str(model)) if model and str(model).strip() else None,
        }
        if _lines(maqsad):
            dars["maqsad"] = _lines(maqsad)
        if _lines(lugat):
            dars["lugat"] = _lines(lugat)
        if _lines(nazariya):
            dars["nazariya"] = [{"title": "Nazariy qism", "points": _lines(nazariya)}]
        if _lines(amaliy):
            dars["amaliy"] = [{"title": "Amaliy qism", "points": _lines(amaliy)}]
        if _lines(uyga):
            dars["uyga"] = _lines(uyga)
        # MET-02 ning uchta maydoni. Boʻsh boʻlsa kalit ham qoʻyilmaydi —
        # JSONB da boʻsh roʻyxat saqlashning maʼnosi yoʻq.
        if natija and str(natija).strip():
            dars["natija"] = _toza(str(natija))
        if _lines(jihoz):
            dars["jihoz"] = _lines(jihoz)
        if _lines(baholash):
            # Kalit `baholash`, `mezon` EMAS: statik Robototexnika
            # bazasida `mezon` boshqa maʼnoda (jadval) band va ikkalasi
            # bir xil koʻrinishda chizilardi.
            dars["baholash"] = _lines(baholash)
        if _lines(resurslar):
            dars["resurslar"] = _lines(resurslar)
        # MET-04: tashqi video havola. Faqat http(s) — `javascript:` yoki
        # `data:` havolasi kartochkaga tushib qolmasin.
        if video and str(video).strip().lower().startswith(("http://", "https://")):
            dars["video"] = str(video).strip()
        elif video and str(video).strip():
            warnings.append(f"{idx}-qator: video havola http(s) bilan boshlansin — tashlandi.")
        lessons.append(dars)

    if not lessons:
        raise ValidationError("Faylda birorta ham dars topilmadi.")
    if len(lessons) > MAX_LESSONS:
        raise ValidationError(f"Darslar soni {MAX_LESSONS} dan oshmasin.")
    # Chorak boʻyicha barqaror tartib (fayl tartibi saqlangan holda).
    lessons.sort(key=lambda d: d["chorak"])
    return lessons, warnings


async def import_plan(
    session: AsyncSession,
    actor: CurrentUser,
    *,
    fan: str,
    yil: str,
    sinf: str,
    data: bytes,
    source_name: str | None,
    ip: str | None = None,
) -> ImportResult:
    fan = _toza(fan)
    if not fan:
        raise ValidationError("Fan nomi boʻsh boʻlmasin.")
    if yil not in PROGRAM_YEARS:
        raise ValidationError("Yil «1-yil» yoki «2-yil» boʻlsin.")
    sinf = sinf.strip()
    if not sinf:
        raise ValidationError("Sinf koʻrsatilsin.")

    lessons, warnings = _parse_xlsx(data)

    plan = CurriculumPlan(
        fan=fan,
        yil=yil,
        sinf=sinf,
        status=CurriculumStatus.QORALAMA.value,
        source_name=(source_name or "").strip()[:200] or None,
        lessons=lessons,
        created_by_id=actor.id,
    )
    session.add(plan)
    await session.flush()
    audit_service.record(
        session,
        object_type="curriculum_plan",
        object_id=plan.id,
        action=AuditAction.CREATE,
        new={"fan": fan, "yil": yil, "sinf": sinf, "darslar": len(lessons)},
        actor_id=actor.id,
        ip=ip,
    )
    await session.commit()
    return ImportResult(plan=plan, warnings=warnings)


# ─────────────────────────── Roʻyxat / joriy ───────────────────────────


async def list_plans(session: AsyncSession) -> list[CurriculumPlan]:
    rows = await session.execute(
        select(CurriculumPlan)
        .where(CurriculumPlan.is_archived.is_(False))
        .order_by(
            CurriculumPlan.fan,
            CurriculumPlan.yil,
            CurriculumPlan.sinf,
            CurriculumPlan.created_at.desc(),
        )
    )
    return list(rows.scalars())


async def get_plan(session: AsyncSession, plan_id: uuid.UUID) -> CurriculumPlan:
    plan = await session.get(CurriculumPlan, plan_id)
    if plan is None or plan.is_archived:
        raise NotFoundError("Reja topilmadi.")
    return plan


async def publish(
    session: AsyncSession,
    actor: CurrentUser,
    plan_id: uuid.UUID,
    *,
    ip: str | None = None,
) -> CurriculumPlan:
    """Rejani JORIY qiladi; shu (fan, yil, sinf)ning eski joriysi ARXIV.

    MET-06: kim joriy qila olishi sozlamaga bogʻliq. Tasdiqlash
    yoqilgan boʻlsa — faqat oʻquv boʻlimi; oʻchirilgan boʻlsa ustoz
    OʻZI yaratgan rejani oʻzi joriy qiladi.

    MET-07: eski versiyani qayta joriy qilish ham shu funksiya —
    «oldingi versiyaga qaytarish» alohida amal emas.
    """
    plan = await get_plan(session, plan_id)
    await assert_can_publish(session, actor)
    if not actor.has(
        RoleName.ACADEMIC.value, RoleName.ADMIN.value, RoleName.SUPERADMIN.value
    ) and plan.created_by_id != actor.id:
        raise PermissionDeniedError("Faqat oʻzingiz qoʻshgan rejani joriy qila olasiz.")
    if plan.status == CurriculumStatus.JORIY.value:
        raise ConflictError("Bu reja allaqachon joriy.")

    eski = (
        await session.execute(
            select(CurriculumPlan).where(
                CurriculumPlan.fan == plan.fan,
                CurriculumPlan.yil == plan.yil,
                CurriculumPlan.sinf == plan.sinf,
                CurriculumPlan.status == CurriculumStatus.JORIY.value,
                CurriculumPlan.is_archived.is_(False),
            )
        )
    ).scalars()
    for e in eski:
        e.status = CurriculumStatus.ARXIV.value

    plan.status = CurriculumStatus.JORIY.value
    audit_service.record(
        session,
        object_type="curriculum_plan",
        object_id=plan.id,
        action=AuditAction.UPDATE,
        old={"status": "qoralama"},
        new={"status": "joriy", "fan": plan.fan, "yil": plan.yil, "sinf": plan.sinf},
        actor_id=actor.id,
        ip=ip,
    )
    await session.commit()
    return plan


async def archive(
    session: AsyncSession, actor: CurrentUser, plan_id: uuid.UUID, *, ip: str | None = None
) -> None:
    plan = await get_plan(session, plan_id)
    plan.is_archived = True
    plan.archived_at = utcnow()
    audit_service.record(
        session,
        object_type="curriculum_plan",
        object_id=plan.id,
        action=AuditAction.ARCHIVE,
        old={"status": plan.status},
        new={"is_archived": True},
        actor_id=actor.id,
        ip=ip,
    )
    await session.commit()


# ─────────────────────────── Eksport ───────────────────────────


def export_xlsx(plan: CurriculumPlan) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reja"
    ws.append([c[0] for c in COLUMNS])
    for d in plan.lessons:
        naz = d.get("nazariya") or []
        ama = d.get("amaliy") or []
        ws.append([
            d.get("chorak", 1),
            d.get("title", ""),
            d.get("type", ""),
            d.get("model") or "",
            "\n".join(d.get("maqsad") or []),
            "\n".join(d.get("lugat") or []),
            "\n".join(pt for b in naz for pt in b.get("points", [])),
            "\n".join(pt for b in ama for pt in b.get("points", [])),
            "\n".join(d.get("uyga") or []),
            d.get("natija") or "",
            "\n".join(d.get("jihoz") or []),
            "\n".join(d.get("baholash") or []),
            "\n".join(d.get("resurslar") or []),
            d.get("video") or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────── Ustozlar uchun joriy rejalar ───────────────────


async def published_catalog(session: AsyncSession) -> dict:
    """fan → yil → sinf → darslar soni (ustoz kabineti fan tanlagichi)."""
    rows = (
        await session.execute(
            select(CurriculumPlan).where(
                CurriculumPlan.status == CurriculumStatus.JORIY.value,
                CurriculumPlan.is_archived.is_(False),
            )
        )
    ).scalars()
    out: dict = {}
    for p in rows:
        out.setdefault(p.fan, {}).setdefault(p.yil, {})[p.sinf] = len(p.lessons)
    return out


async def published_plan(
    session: AsyncSession, *, fan: str, yil: str, sinf: str
) -> CurriculumPlan:
    plan = await session.scalar(
        select(CurriculumPlan).where(
            CurriculumPlan.fan == fan,
            CurriculumPlan.yil == yil,
            CurriculumPlan.sinf == sinf,
            CurriculumPlan.status == CurriculumStatus.JORIY.value,
            CurriculumPlan.is_archived.is_(False),
        )
    )
    if plan is None:
        raise NotFoundError("Bu fan/sinf uchun joriy reja yoʻq.")
    return plan


# ─────────────────── Qidiruv (MET-05) ───────────────────


@dataclass(frozen=True, slots=True)
class SearchHit:
    plan_id: uuid.UUID
    fan: str
    yil: str
    sinf: str
    chorak: int
    #: Reja ichidagi dars tartibi — kartochkani ochish uchun.
    index: int
    title: str
    #: Nima boʻyicha topildi: «mavzu», «atama» yoki «jihoz».
    matched_in: str


def _matches(dars: dict, q: str) -> str | None:
    """Dars qidiruv soʻziga mos keladimi va NIMA boʻyicha.

    TZ MET-05 uchta manbani nomlaydi: mavzu nomi, atama va jihoz nomi.
    Tartib muhim — foydalanuvchi «nima uchun topildi» degan savolga
    javob koʻradi, shuning uchun eng aniq mos birinchi tekshiriladi.
    """
    if q in (dars.get("title") or "").casefold():
        return "mavzu"
    for atama in dars.get("lugat") or []:
        if q in str(atama).casefold():
            return "atama"
    for jihoz in dars.get("jihoz") or []:
        if q in str(jihoz).casefold():
            return "jihoz"
    return None


async def search_lessons(
    session: AsyncSession,
    *,
    q: str,
    fan: str | None = None,
    sinf: str | None = None,
    chorak: int | None = None,
    limit: int = 50,
) -> list[SearchHit]:
    """Joriy rejalar boʻyicha qidiruv va filtr (MET-05).

    Filtrlash SQL da, matn qidiruvi esa PYTHON da. Sabab: darslar
    JSONB roʻyxatida yotadi va bitta (fan, yil, sinf) uchun bitta
    joriy reja boʻladi — maktab hajmida bu bir necha oʻnlab qator,
    yaʼni bir necha ming dars. GIN indeks va `jsonb_path_query` bu
    hajmda foyda bermaydi, kodni esa sezilarli murakkablashtiradi.
    Rejalar soni yuzlab boʻlsa — oʻshanda indeks qoʻyiladi.
    """
    soz = _toza(q).casefold()
    if len(soz) < 2:
        raise ValidationError("Qidiruv soʻzi kamida ikki belgidan iborat boʻlsin.")

    stmt = select(CurriculumPlan).where(
        CurriculumPlan.status == CurriculumStatus.JORIY.value,
        CurriculumPlan.is_archived.is_(False),
    )
    if fan:
        stmt = stmt.where(CurriculumPlan.fan == _toza(fan))
    if sinf:
        stmt = stmt.where(CurriculumPlan.sinf == sinf.strip())

    natija: list[SearchHit] = []
    for plan in (await session.execute(stmt)).scalars():
        for i, dars in enumerate(plan.lessons):
            if chorak is not None and dars.get("chorak") != chorak:
                continue
            qayerda = _matches(dars, soz)
            if qayerda is None:
                continue
            natija.append(
                SearchHit(
                    plan_id=plan.id,
                    fan=plan.fan,
                    yil=plan.yil,
                    sinf=plan.sinf,
                    chorak=int(dars.get("chorak") or 1),
                    index=i,
                    title=str(dars.get("title") or ""),
                    matched_in=qayerda,
                )
            )
            if len(natija) >= limit:
                return natija
    return natija


# ─────────── Ustoz rejasi va versiyalar (MET-06, MET-07) ───────────


async def _requires_approval(session: AsyncSession) -> bool:
    """Sozlama yoqilganmi. Sozlama qatori yoʻq boʻlsa — YOQILGAN.

    Sukut xavfsiz tomonga: sozlama toʻldirilmagan maktabda ustoz
    rejasi tekshiruvsiz nashr boʻlib ketmasin.
    """
    row = await session.scalar(
        select(SchoolSettings).where(SchoolSettings.is_archived.is_(False))
    )
    return True if row is None else row.curriculum_requires_approval


async def create_plan(
    session: AsyncSession,
    actor: CurrentUser,
    *,
    fan: str,
    yil: str,
    sinf: str,
    lessons: list[dict],
    ip: str | None = None,
) -> CurriculumPlan:
    """MET-06: ustoz oʻz dars rejasini qoʻshadi.

    Reja har doim QORALAMA boʻlib tugʻiladi. Uni joriy qilish alohida
    amal (`publish`) va u tasdiqlash sozlamasiga bogʻliq: yoqilgan
    boʻlsa faqat oʻquv boʻlimi joriy qiladi.
    """
    fan = _toza(fan)
    if not fan:
        raise ValidationError("Fan nomi boʻsh boʻlmasin.")
    if yil not in PROGRAM_YEARS:
        raise ValidationError("Yil «1-yil» yoki «2-yil» boʻlsin.")
    sinf = sinf.strip()
    if not sinf:
        raise ValidationError("Sinfni koʻrsating.")
    if not lessons:
        raise ValidationError("Rejada kamida bitta dars boʻlsin.")
    if len(lessons) > MAX_LESSONS:
        raise ValidationError(f"Darslar soni {MAX_LESSONS} dan oshmasin.")

    plan = CurriculumPlan(
        fan=fan,
        yil=yil,
        sinf=sinf,
        lessons=lessons,
        status=CurriculumStatus.QORALAMA.value,
        source_name=None,
        created_by_id=actor.id,
    )
    session.add(plan)
    await session.flush()

    audit_service.record(
        session,
        object_type="curriculum_plan",
        object_id=plan.id,
        action=AuditAction.CREATE,
        new={"fan": fan, "yil": yil, "sinf": sinf, "darslar": len(lessons), "manba": "qoʻlda"},
        actor_id=actor.id,
        ip=ip,
    )
    await session.commit()
    await session.refresh(plan)
    return plan


async def assert_can_publish(session: AsyncSession, actor: CurrentUser) -> None:
    """Kim joriy qila oladi (MET-06).

    Sozlama YOQILGAN boʻlsa — faqat oʻquv boʻlimi/administrator.
    Oʻchirilgan boʻlsa ustoz ham oʻzi joriy qiladi.
    """
    if actor.has(
        RoleName.ACADEMIC.value, RoleName.ADMIN.value, RoleName.SUPERADMIN.value
    ):
        return
    if await _requires_approval(session):
        raise PermissionDeniedError(
            "Reja nashr etilishidan oldin oʻquv boʻlimi tasdiqlashi kerak."
        )
    if not actor.is_teacher:
        raise PermissionDeniedError("Reja joriy qilishga ruxsatingiz yoʻq.")


async def list_versions(
    session: AsyncSession, *, fan: str, yil: str, sinf: str
) -> list[CurriculumPlan]:
    """MET-07: shu (fan, yil, sinf) uchun barcha versiyalar.

    Eskisi OʻCHIRILMAYDI — joriy qilinganda `arxiv` holatiga oʻtadi
    (CLAUDE.md 1-qoida). Shu sabab «oldingi versiyaga qaytarish»
    alohida amal emas: eski versiyani qayta joriy qilish yetadi.
    """
    rows = await session.execute(
        select(CurriculumPlan)
        .where(
            CurriculumPlan.fan == _toza(fan),
            CurriculumPlan.yil == yil,
            CurriculumPlan.sinf == sinf.strip(),
            CurriculumPlan.is_archived.is_(False),
        )
        .order_by(CurriculumPlan.created_at.desc())
    )
    return list(rows.scalars())


#: Kartochkada tahrirlanadigan matn maydonlari (MET-02).
_TEXT_FIELDS = ("title", "model", "natija", "video")
#: Roʻyxat koʻrinishidagi maydonlar.
_LIST_FIELDS = ("maqsad", "lugat", "jihoz", "baholash", "uyga", "resurslar")


async def _assert_can_edit_plan(
    session: AsyncSession, actor: CurrentUser, plan: CurriculumPlan
) -> None:
    """Kartochkani kim tahrirlaydi.

    Oʻquv boʻlimi va administrator — har doim. Ustoz esa faqat OʻZI
    yaratgan va hali QORALAMA turgan rejani: joriy qilingan reja butun
    maktabga tarqalgan hujjat, uni jimgina oʻzgartirib boʻlmaydi.
    """
    if actor.has(
        RoleName.ACADEMIC.value, RoleName.ADMIN.value, RoleName.SUPERADMIN.value
    ):
        return
    if plan.created_by_id == actor.id and plan.status == CurriculumStatus.QORALAMA.value:
        return
    raise PermissionDeniedError("Bu rejani tahrirlashga ruxsatingiz yoʻq.")


async def update_lesson_card(
    session: AsyncSession,
    actor: CurrentUser,
    *,
    plan_id: uuid.UUID,
    index: int,
    changes: dict,
    ip: str | None = None,
) -> dict:
    """Bitta dars kartochkasini tahrirlaydi (MET-02, MET-03, MET-04).

    Faqat YUBORILGAN maydonlar oʻzgaradi — kartochkaning qolgan qismi
    tegilmaydi. Fayl ilovasi `files` roʻyxatida `{id, name}` boʻlib
    turadi; baytlar `stored_files` da (CLAUDE.md 10-qoida).
    """
    plan = await get_plan(session, plan_id)
    await _assert_can_edit_plan(session, actor, plan)

    if index < 0 or index >= len(plan.lessons):
        raise NotFoundError("Dars kartochkasi topilmadi.")

    # JSONB ustuni oʻrnida oʻzgartirilsa SQLAlchemy uni «iflos» deb
    # koʻrmaydi — roʻyxat butunlay qayta yoziladi.
    darslar = [dict(d) for d in plan.lessons]
    dars = darslar[index]
    eski = {k: dars.get(k) for k in changes if k in dars}

    for nom in _TEXT_FIELDS:
        if nom not in changes:
            continue
        qiymat = changes[nom]
        if nom == "video" and qiymat:
            if not str(qiymat).lower().startswith(("http://", "https://")):
                raise ValidationError("Video havola http:// yoki https:// bilan boshlansin.")
        dars[nom] = _toza(str(qiymat)) if qiymat else None

    for nom in _LIST_FIELDS:
        if nom not in changes:
            continue
        qatorlar = [_toza(str(x)) for x in (changes[nom] or []) if str(x).strip()]
        dars[nom] = qatorlar

    if "files" in changes:
        # Faylning oʻzi tekshiriladi: mavjud boʻlmagan id yozilsa
        # kartochkada «ochilmaydigan ilova» qolib ketardi.
        ilovalar = []
        for f in changes["files"] or []:
            fayl = await storage.get(session, uuid.UUID(str(f["id"])))
            ilovalar.append({"id": str(fayl.id), "name": fayl.original_name})
        dars["files"] = ilovalar

    plan.lessons = darslar
    audit_service.record(
        session,
        object_type="curriculum_plan",
        object_id=plan.id,
        action=AuditAction.UPDATE,
        old={"index": index, **eski},
        new={"index": index, **{k: dars.get(k) for k in changes}},
        actor_id=actor.id,
        ip=ip,
    )
    await session.commit()
    return dars
