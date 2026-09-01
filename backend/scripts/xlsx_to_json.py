"""Google Sheets xlsx eksportini import_real.py uchun JSON'ga aylantiradi.

Ishlatish: python scripts/xlsx_to_json.py real_data.xlsx real_data.json
(openpyxl kerak — tizim pythonida bor.)
"""

import json
import sys
from pathlib import Path

import openpyxl


def main() -> None:
    src, dst = Path(sys.argv[1]), Path(sys.argv[2])
    wb = openpyxl.load_workbook(src, read_only=True)

    teachers = []
    for row in wb.worksheets[0].iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue
        teachers.append(
            {
                "full": str(row[1]),
                "subject": str(row[2]) if row[2] else None,
                "class": str(row[3]) if len(row) > 3 and row[3] else None,
            }
        )

    students = []
    for ws in wb.worksheets[1:]:
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[1]:
                continue
            students.append(
                {
                    "full": str(row[1]),
                    "class": ws.title,
                    "phone": str(row[3]) if len(row) > 3 and row[3] else "",
                }
            )

    dst.write_text(
        json.dumps({"teachers": teachers, "students": students}, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )
    print(f"{len(teachers)} ustoz, {len(students)} o'quvchi -> {dst}")


if __name__ == "__main__":
    main()
