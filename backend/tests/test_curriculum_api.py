"""Oʻquv rejalari (metodik baza) — import/joriy/eksport oqimi."""

import io

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models import Role, RoleName, SchoolSettings, User

PASSWORD = "Sinov12345!"  # noqa: S106


async def _roles(session: AsyncSession) -> dict[str, Role]:
    return {r.name: r for r in (await session.execute(select(Role))).scalars()}


async def _user(session, roles, names, login, last):  # noqa: ANN001, ANN202
    u = User(
        login=login, password_hash=hash_password(PASSWORD), last_name=last, first_name="Sinov"
    )
    u.roles = [roles[r] for r in names]
    session.add(u)
    await session.flush()
    return u


@pytest.fixture
async def world(session: AsyncSession) -> dict:
    roles = await _roles(session)
    oquv = await _user(session, roles, [RoleName.ACADEMIC.value], "cu.oquv", "Oquvboshi")
    ustoz = await _user(session, roles, [RoleName.TEACHER.value], "cu.ustoz", "Aliyev")
    ota = await _user(session, roles, [RoleName.PARENT.value], "cu.ota", "Karimov")
    return {"oquv": oquv, "ustoz": ustoz, "ota": ota}


async def _token(client: AsyncClient, login: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"login": login, "password": PASSWORD})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _auth(t: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {t}"}


def _xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reja"
    ws.append([
        "Chorak", "Mavzu", "Tur", "Model", "Maqsad", "Lugʻat",
        "Nazariya", "Amaliy", "Uyga vazifa", "Resurslar",
    ])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _import(client, token, fan="Matematika", yil="1-yil", sinf="1-A", rows=None):  # noqa: ANN001, ANN202
    data = _xlsx(
        rows
        if rows is not None
        else [
            [
                1, "Sonlar bilan tanishuv", "qurish", "",
                "Maqsad 1\nMaqsad 2", "Son (Number) — miqdor",
                "Kirish", "Mashq", "5 ta misol", "Doska",
            ],
            [2, "Qo'shish amali", "dasturlash", "", "Maqsad", "", "", "", "", ""],
        ]
    )
    return await client.post(
        "/api/v1/curriculum/import",
        headers=_auth(token),
        data={"fan": fan, "yil": yil, "sinf": sinf},
        files={
            "file": (
                "reja.xlsx",
                data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


async def test_shablon_yuklanadi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    r = await client.get("/api/v1/curriculum/template", headers=_auth(t))
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "Reja" in wb.sheetnames and "Yoʻriqnoma" in wb.sheetnames


async def test_import_va_joriy_oqimi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    r = await _import(client, t)
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["plan"]["status"] == "qoralama"
    assert body["plan"]["darslar_soni"] == 2
    plan_id = body["plan"]["id"]

    # Apostrof normalizatsiyasi: Qo'shish -> Qoʻshish
    r = await client.get(f"/api/v1/curriculum/plans/{plan_id}", headers=_auth(t))
    darslar = r.json()["lessons"]
    assert any("Qoʻshish" in d["title"] for d in darslar)

    # Joriy qilinmaguncha ustozga ko'rinmaydi
    ustoz = await _token(client, "cu.ustoz")
    r = await client.get("/api/v1/curriculum/published", headers=_auth(ustoz))
    assert r.json()["fanlar"] == {}

    # Joriy qilish
    r = await client.post(f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(t))
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "joriy"

    # Endi ustoz katalog va rejani oladi
    r = await client.get("/api/v1/curriculum/published", headers=_auth(ustoz))
    assert r.json()["fanlar"]["Matematika"]["1-yil"]["1-A"] == 2
    r = await client.get(
        "/api/v1/curriculum/published/plan",
        headers=_auth(ustoz),
        params={"fan": "Matematika", "yil": "1-yil", "sinf": "1-A"},
    )
    assert r.status_code == 200
    assert len(r.json()["lessons"]) == 2


async def test_yangi_joriy_eskisini_arxivlaydi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    r1 = (await _import(client, t)).json()["plan"]["id"]
    await client.post(f"/api/v1/curriculum/plans/{r1}/publish", headers=_auth(t))
    r2 = (await _import(client, t)).json()["plan"]["id"]
    await client.post(f"/api/v1/curriculum/plans/{r2}/publish", headers=_auth(t))

    r = await client.get("/api/v1/curriculum/plans", headers=_auth(t))
    holatlar = {p["id"]: p["status"] for p in r.json()}
    assert holatlar[r1] == "arxiv"
    assert holatlar[r2] == "joriy"


async def test_import_ogohlantirishlari(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    r = await _import(
        client,
        t,
        rows=[
            [9, "Chorak xato", "yoq-tur", "", "", "", "", "", "", ""],
            [None, "", "", "", "", "", "", "", "", ""],
            [2, "To'g'ri dars", "qurish", "", "", "", "", "", "", ""],
        ],
    )
    assert r.status_code == 201, r.text
    warnings = r.json()["warnings"]
    assert any("chorak" in w for w in warnings)
    assert any("tur" in w for w in warnings)


async def test_eksport_ishlaydi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    plan_id = (await _import(client, t)).json()["plan"]["id"]
    r = await client.get(f"/api/v1/curriculum/plans/{plan_id}/export", headers=_auth(t))
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb["Reja"].max_row == 3  # sarlavha + 2 dars


async def test_ustoz_va_otaona_boshqara_olmaydi(client: AsyncClient, world: dict) -> None:
    """Rol darvozasi: import/joriy faqat o'quv bo'limi/admin."""
    ustoz = await _token(client, "cu.ustoz")
    r = await _import(client, ustoz)
    assert r.status_code == 403, r.text
    assert (
        await client.get("/api/v1/curriculum/plans", headers=_auth(ustoz))
    ).status_code == 403

    # Ota-ona joriy katalogni ham ko'rmasin degan talab yo'q — lekin
    # boshqaruvga kira olmasligi shart.
    ota = await _token(client, "cu.ota")
    assert (
        await client.get("/api/v1/curriculum/template", headers=_auth(ota))
    ).status_code == 403


# ─────────── Yangi maydonlar, qidiruv, versiyalar (MET-02…MET-07) ───────────


def _xlsx_toliq(rows: list[list]) -> bytes:
    """Toʻliq ustunli shablon — MET-02 ning uchta yangi maydoni bilan."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reja"
    ws.append([
        "Chorak", "Mavzu", "Tur", "Model", "Maqsad", "Lugʻat",
        "Nazariya", "Amaliy", "Uyga vazifa",
        "Kutilayotgan natija", "Kerakli jihozlar", "Baholash mezoni",
        "Resurslar", "Video havola",
    ])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _import_toliq(client, token, **kw):  # noqa: ANN001, ANN202
    rows = kw.pop("rows", None) or [
        [
            1, "Kuchlanish va tok", "elektronika", "",
            "Oʻlchashni oʻrganish", "Volt (Volt) — kuchlanish birligi",
            "Nazariya", "Amaliy",
            "Uyga", "Oʻquvchi multimetrdan foydalana oladi",
            "Multimetr\nBatareyka", "Oʻlchash aniqligi",
            "Darslik", "https://youtu.be/abc",
        ],
    ]
    return await client.post(
        "/api/v1/curriculum/import",
        headers=_auth(token),
        data={
            "fan": kw.get("fan", "Fizika"),
            "yil": kw.get("yil", "1-yil"),
            "sinf": kw.get("sinf", "5-A"),
        },
        files={
            "file": (
                "reja.xlsx",
                _xlsx_toliq(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


async def test_yangi_maydonlar_import_qilinadi(client: AsyncClient, world: dict) -> None:
    """MET-02: kutilayotgan natija, jihozlar, baholash mezoni. MET-04: video."""
    t = await _token(client, "cu.oquv")
    r = await _import_toliq(client, t)
    assert r.status_code == 201, r.text

    plan_id = r.json()["plan"]["id"]
    detail = await client.get(f"/api/v1/curriculum/plans/{plan_id}", headers=_auth(t))
    dars = detail.json()["lessons"][0]
    assert dars["natija"] == "Oʻquvchi multimetrdan foydalana oladi"
    assert dars["jihoz"] == ["Multimetr", "Batareyka"]
    assert dars["baholash"] == ["Oʻlchash aniqligi"]
    assert dars["video"] == "https://youtu.be/abc"


async def test_xavfli_video_havola_tashlanadi(client: AsyncClient, world: dict) -> None:
    """`javascript:` havolasi kartochkaga tushib qolmasin."""
    t = await _token(client, "cu.oquv")
    r = await _import_toliq(
        client,
        t,
        rows=[
            [
                1, "Mavzu", "qurish", "", "", "", "", "", "",
                "", "", "", "", "javascript:alert(1)",
            ]
        ],
    )
    assert r.status_code == 201
    assert any("video havola" in w for w in r.json()["warnings"])

    plan_id = r.json()["plan"]["id"]
    detail = await client.get(f"/api/v1/curriculum/plans/{plan_id}", headers=_auth(t))
    assert "video" not in detail.json()["lessons"][0]


async def test_qidiruv_mavzu_atama_va_jihoz_boyicha(
    client: AsyncClient, world: dict
) -> None:
    """MET-05 uchta manbani nomlaydi va natija qaysi biri ekanini aytadi."""
    t = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, t)).json()["plan"]["id"]
    await client.post(f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(t))

    ustoz = await _token(client, "cu.ustoz")
    for soz, qayerda in [
        ("kuchlanish", "mavzu"),
        ("volt", "atama"),
        ("multimetr", "jihoz"),
    ]:
        r = await client.get(
            "/api/v1/curriculum/search", headers=_auth(ustoz), params={"q": soz}
        )
        assert r.status_code == 200, r.text
        rows = r.json()
        assert len(rows) == 1, soz
        assert rows[0]["matched_in"] == qayerda


async def test_qidiruv_qoralamani_topmaydi(client: AsyncClient, world: dict) -> None:
    """Qoralama hali hujjat emas."""
    t = await _token(client, "cu.oquv")
    await _import_toliq(client, t)

    r = await client.get(
        "/api/v1/curriculum/search", headers=_auth(t), params={"q": "kuchlanish"}
    )
    assert r.json() == []


async def test_qidiruvda_chorak_filtri(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, t)).json()["plan"]["id"]
    await client.post(f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(t))

    ustoz = await _token(client, "cu.ustoz")
    topildi = await client.get(
        "/api/v1/curriculum/search",
        headers=_auth(ustoz),
        params={"q": "kuchlanish", "chorak": 1},
    )
    assert len(topildi.json()) == 1

    yoq = await client.get(
        "/api/v1/curriculum/search",
        headers=_auth(ustoz),
        params={"q": "kuchlanish", "chorak": 3},
    )
    assert yoq.json() == []


async def test_ustoz_oz_rejasini_qoshadi_lekin_joriy_qila_olmaydi(
    client: AsyncClient, world: dict
) -> None:
    """MET-06: tasdiqlash sukut boʻyicha YOQILGAN."""
    ustoz = await _token(client, "cu.ustoz")
    r = await client.post(
        "/api/v1/curriculum/plans",
        headers=_auth(ustoz),
        json={
            "fan": "Informatika",
            "yil": "1-yil",
            "sinf": "6-A",
            "lessons": [{"chorak": 1, "title": "Algoritm", "type": "dasturlash"}],
        },
    )
    assert r.status_code == 201, r.text
    assert r.json()["status"] == "qoralama"

    plan_id = r.json()["id"]
    joriy = await client.post(
        f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(ustoz)
    )
    assert joriy.status_code == 403


async def test_oquv_bolimi_ustoz_rejasini_tasdiqlaydi(
    client: AsyncClient, world: dict
) -> None:
    ustoz = await _token(client, "cu.ustoz")
    plan_id = (
        await client.post(
            "/api/v1/curriculum/plans",
            headers=_auth(ustoz),
            json={
                "fan": "Informatika",
                "yil": "1-yil",
                "sinf": "6-A",
                "lessons": [{"chorak": 1, "title": "Algoritm", "type": "dasturlash"}],
            },
        )
    ).json()["id"]

    oquv = await _token(client, "cu.oquv")
    r = await client.post(
        f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(oquv)
    )
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "joriy"


async def test_sozlama_ochirilsa_ustoz_ozi_joriy_qiladi(
    client: AsyncClient, world: dict, session: AsyncSession
) -> None:
    """MET-06 ning ikkinchi yarmi — sozlama oʻchirilgan holat."""
    session.add(SchoolSettings(name="Tarbion", curriculum_requires_approval=False))
    await session.commit()

    ustoz = await _token(client, "cu.ustoz")
    plan_id = (
        await client.post(
            "/api/v1/curriculum/plans",
            headers=_auth(ustoz),
            json={
                "fan": "Informatika",
                "yil": "1-yil",
                "sinf": "6-A",
                "lessons": [{"chorak": 1, "title": "Algoritm", "type": "dasturlash"}],
            },
        )
    ).json()["id"]

    r = await client.post(
        f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(ustoz)
    )
    assert r.status_code == 200, r.text


async def test_ustoz_begona_rejani_joriy_qila_olmaydi(
    client: AsyncClient, world: dict, session: AsyncSession
) -> None:
    session.add(SchoolSettings(name="Tarbion", curriculum_requires_approval=False))
    await session.commit()

    oquv = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, oquv)).json()["plan"]["id"]

    ustoz = await _token(client, "cu.ustoz")
    r = await client.post(
        f"/api/v1/curriculum/plans/{plan_id}/publish", headers=_auth(ustoz)
    )
    assert r.status_code == 403


async def test_versiyalar_saqlanadi_va_eskisiga_qaytariladi(
    client: AsyncClient, world: dict
) -> None:
    """MET-07: eski versiya oʻchirilmaydi, qayta joriy qilinadi."""
    t = await _token(client, "cu.oquv")

    birinchi = (await _import_toliq(client, t)).json()["plan"]["id"]
    await client.post(f"/api/v1/curriculum/plans/{birinchi}/publish", headers=_auth(t))

    ikkinchi = (
        await _import_toliq(
            client,
            t,
            rows=[[2, "Qarshilik", "elektronika", "", "", "", "", "", "", "", "", "", "", ""]],
        )
    ).json()["plan"]["id"]
    await client.post(f"/api/v1/curriculum/plans/{ikkinchi}/publish", headers=_auth(t))

    r = await client.get(
        "/api/v1/curriculum/versions",
        headers=_auth(t),
        params={"fan": "Fizika", "yil": "1-yil", "sinf": "5-A"},
    )
    assert r.status_code == 200, r.text
    versiyalar = {v["id"]: v["status"] for v in r.json()}
    assert versiyalar[birinchi] == "arxiv"
    assert versiyalar[ikkinchi] == "joriy"

    # Eskisiga qaytarish — oʻsha versiyani qayta joriy qilish.
    await client.post(f"/api/v1/curriculum/plans/{birinchi}/publish", headers=_auth(t))
    r2 = await client.get(
        "/api/v1/curriculum/versions",
        headers=_auth(t),
        params={"fan": "Fizika", "yil": "1-yil", "sinf": "5-A"},
    )
    yangi = {v["id"]: v["status"] for v in r2.json()}
    assert yangi[birinchi] == "joriy"
    assert yangi[ikkinchi] == "arxiv"


async def test_kartochka_tahrirlanadi(client: AsyncClient, world: dict) -> None:
    """MET-02: faqat yuborilgan maydon oʻzgaradi, qolgani tegilmaydi."""
    t = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, t)).json()["plan"]["id"]

    r = await client.patch(
        f"/api/v1/curriculum/plans/{plan_id}/lessons/0",
        headers=_auth(t),
        json={"jihoz": ["Ossillograf"], "natija": "Yangi natija"},
    )
    assert r.status_code == 200, r.text
    dars = r.json()
    assert dars["jihoz"] == ["Ossillograf"]
    assert dars["natija"] == "Yangi natija"
    # Tegilmagan maydon joyida.
    assert dars["title"] == "Kuchlanish va tok"


async def test_kartochkaga_xavfli_video_qoyilmaydi(
    client: AsyncClient, world: dict
) -> None:
    t = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, t)).json()["plan"]["id"]

    r = await client.patch(
        f"/api/v1/curriculum/plans/{plan_id}/lessons/0",
        headers=_auth(t),
        json={"video": "javascript:alert(1)"},
    )
    assert r.status_code == 422


async def test_ustoz_begona_kartochkani_tahrirlay_olmaydi(
    client: AsyncClient, world: dict
) -> None:
    """Joriy reja butun maktabga tarqalgan hujjat — `403`, `404` emas (X-3)."""
    t = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, t)).json()["plan"]["id"]

    ustoz = await _token(client, "cu.ustoz")
    r = await client.patch(
        f"/api/v1/curriculum/plans/{plan_id}/lessons/0",
        headers=_auth(ustoz),
        json={"title": "Oʻzgartirdim"},
    )
    assert r.status_code == 403


async def test_mavjud_bolmagan_kartochka_404(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "cu.oquv")
    plan_id = (await _import_toliq(client, t)).json()["plan"]["id"]

    r = await client.patch(
        f"/api/v1/curriculum/plans/{plan_id}/lessons/99",
        headers=_auth(t),
        json={"title": "Yoʻq"},
    )
    assert r.status_code == 404
