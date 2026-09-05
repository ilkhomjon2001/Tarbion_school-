"""Testlar moduli (TST-01…TST-05).

Eng muhim testlar:
  · toʻgʻri javob oʻquvchiga HECH QACHON yuborilmaydi
  · ball serverda hisoblanadi — frontend yubormaydi
  · urinishlar soni serverda cheklanadi
  · boshqa sinfning testini ocha olmaydi (X-1)
  · savol ham, test ham OʻCHMAYDI — arxivlanadi (1-qoida)
"""

import io
from datetime import date, timedelta

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.core.timeutil import utcnow
from app.models import (
    AcademicYear,
    Guardian,
    Role,
    RoleName,
    ScheduleEntry,
    SchoolClass,
    Student,
    Subject,
    Test,
    TestAttempt,
    TestQuestion,
    User,
)

PASSWORD = "Sinov12345!"  # noqa: S106


async def _roles(session: AsyncSession) -> dict[str, Role]:
    return {r.name: r for r in (await session.execute(select(Role))).scalars()}


async def _user(
    session: AsyncSession, roles: dict[str, Role], names: list[str], login: str, last: str
) -> User:
    u = User(
        login=login, password_hash=hash_password(PASSWORD), last_name=last, first_name="Sinov"
    )
    u.roles = [roles[r] for r in names]
    session.add(u)
    await session.flush()
    return u


@pytest.fixture
async def world(session: AsyncSession) -> dict[str, object]:
    roles = await _roles(session)

    ustoz = await _user(session, roles, [RoleName.TEACHER.value], "ts.ustoz", "Aliyev")
    begona = await _user(session, roles, [RoleName.TEACHER.value], "ts.begona", "Valiyev")
    ota_a = await _user(session, roles, [RoleName.PARENT.value], "ts.ota_a", "Karimov")
    ota_b = await _user(session, roles, [RoleName.PARENT.value], "ts.ota_b", "Rahimov")
    ali_u = await _user(session, roles, [RoleName.STUDENT.value], "ts.ali", "Abdullayev")
    vali_u = await _user(session, roles, [RoleName.STUDENT.value], "ts.vali", "Boboyev")

    year = AcademicYear(
        name="2026-2027", starts_on=date(2026, 8, 1), ends_on=date(2027, 5, 25), is_current=True
    )
    session.add(year)
    await session.flush()

    math = Subject(name="Matematika", short_name="Mat")
    fizika = Subject(name="Fizika", short_name="Fiz")
    session.add_all([math, fizika])
    await session.flush()

    cls_a = SchoolClass(academic_year_id=year.id, name="8-A")
    cls_b = SchoolClass(academic_year_id=year.id, name="8-B")
    session.add_all([cls_a, cls_b])
    await session.flush()

    ali = Student(class_id=cls_a.id, last_name="Abdullayev", first_name="Ali", user_id=ali_u.id)
    # Boshqa SINFdagi oʻquvchi — testni koʻrmasligi kerak.
    vali = Student(class_id=cls_b.id, last_name="Boboyev", first_name="Vali", user_id=vali_u.id)
    session.add_all([ali, vali])
    await session.flush()

    session.add_all(
        [
            Guardian(student_id=ali.id, user_id=ota_a.id, relation="father"),
            Guardian(student_id=vali.id, user_id=ota_b.id, relation="father"),
            ScheduleEntry(
                academic_year_id=year.id,
                class_id=cls_a.id,
                subject_id=math.id,
                teacher_id=ustoz.id,
                weekday=1,
                period=1,
            ),
        ]
    )
    await session.flush()

    return {
        "ustoz": ustoz,
        "begona": begona,
        "ota_a": ota_a,
        "ota_b": ota_b,
        "class_a": cls_a,
        "class_b": cls_b,
        "math": math,
        "fizika": fizika,
        "ali": ali,
        "vali": vali,
    }


async def _token(client: AsyncClient, login: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"login": login, "password": PASSWORD})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _auth(t: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {t}"}


async def _create_test(client: AsyncClient, token: str, world: dict, **farq: object) -> dict:
    body = {
        "class_id": str(world["class_a"].id),
        "subject_id": str(world["math"].id),
        "title": "Kvadrat tenglamalar",
        "duration_minutes": 30,
        "attempts_allowed": 1,
        "opens_at": (utcnow() - timedelta(hours=1)).isoformat(),
        "closes_at": (utcnow() + timedelta(days=2)).isoformat(),
        "shuffle": False,
    }
    body.update(farq)
    r = await client.post("/api/v1/tests", headers=_auth(token), json=body)
    assert r.status_code == 201, r.text
    return r.json()


async def _add_question(
    client: AsyncClient, token: str, test_id: str, *, correct: int = 0, points: int = 1
) -> dict:
    r = await client.post(
        f"/api/v1/tests/{test_id}/questions",
        headers=_auth(token),
        json={
            "text": "x² − 5x + 6 = 0 ildizlari?",
            "kind": "single",
            "points": points,
            "options": [
                {"text": "x = 2 va x = 3", "is_correct": correct == 0},
                {"text": "x = 1 va x = 6", "is_correct": correct == 1},
                {"text": "Ildizi yoʻq", "is_correct": correct == 2},
            ],
        },
    )
    assert r.status_code == 201, r.text
    return r.json()


async def _publish(client: AsyncClient, token: str, test_id: str) -> None:
    r = await client.put(
        f"/api/v1/tests/{test_id}/status", headers=_auth(token), json={"status": "published"}
    )
    assert r.status_code == 200, r.text


# ─────────────────────────── Test tuzish ───────────────────────────


async def test_test_yaratiladi_qoralama_holatida(client: AsyncClient, world: dict) -> None:
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)

    assert t["status"] == "draft"
    assert t["class_name"] == "8-A"
    assert t["subject_name"] == "Matematika"
    assert t["question_count"] == 0
    assert t["total_students"] == 1


async def test_ozga_fandan_test_tuza_olmaydi(client: AsyncClient, world: dict) -> None:
    """Jurnal bilan bir xil qoida: oʻz sinfi va oʻz fani."""
    token = await _token(client, "ts.ustoz")
    r = await client.post(
        "/api/v1/tests",
        headers=_auth(token),
        json={
            "class_id": str(world["class_a"].id),
            "subject_id": str(world["fizika"].id),
            "title": "Fizika testi",
            "opens_at": utcnow().isoformat(),
            "closes_at": (utcnow() + timedelta(days=1)).isoformat(),
        },
    )
    assert r.status_code == 403, r.text


async def test_savolsiz_test_elon_qilinmaydi(client: AsyncClient, world: dict) -> None:
    """Oʻquvchi boʻsh ekran koʻrmasin."""
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)

    r = await client.put(
        f"/api/v1/tests/{t['id']}/status", headers=_auth(token), json={"status": "published"}
    )
    assert r.status_code == 422, r.text


async def test_hamma_variant_togri_bolsa_rad_etiladi(client: AsyncClient, world: dict) -> None:
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)

    r = await client.post(
        f"/api/v1/tests/{t['id']}/questions",
        headers=_auth(token),
        json={
            "text": "Savol",
            "kind": "single",
            "points": 1,
            "options": [
                {"text": "A", "is_correct": True},
                {"text": "B", "is_correct": True},
            ],
        },
    )
    assert r.status_code == 422, r.text


async def test_single_turida_ikkita_togri_javob_bolmaydi(
    client: AsyncClient, world: dict
) -> None:
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)

    r = await client.post(
        f"/api/v1/tests/{t['id']}/questions",
        headers=_auth(token),
        json={
            "text": "Savol",
            "kind": "single",
            "points": 1,
            "options": [
                {"text": "A", "is_correct": True},
                {"text": "B", "is_correct": True},
                {"text": "C", "is_correct": False},
            ],
        },
    )
    assert r.status_code == 422, r.text


async def test_elon_qilingandan_keyin_savol_qoshilmaydi(
    client: AsyncClient, world: dict
) -> None:
    """Bir xil testni ikki oʻquvchi ikki xil koʻrmasin."""
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)
    await _add_question(client, token, t["id"])
    await _publish(client, token, t["id"])

    r = await client.post(
        f"/api/v1/tests/{t['id']}/questions",
        headers=_auth(token),
        json={
            "text": "Yangi savol",
            "kind": "single",
            "points": 1,
            "options": [
                {"text": "A", "is_correct": True},
                {"text": "B", "is_correct": False},
            ],
        },
    )
    assert r.status_code == 409, r.text


async def test_savol_arxivlanadi(
    client: AsyncClient, world: dict, session: AsyncSession
) -> None:
    """CLAUDE.md 1-qoida."""
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)
    q = await _add_question(client, token, t["id"])

    r = await client.post(
        f"/api/v1/tests/questions/{q['id']}/archive", headers=_auth(token)
    )
    assert r.status_code == 204

    barchasi = (await session.execute(select(TestQuestion))).scalars().all()
    assert len(barchasi) == 1, "savol oʻchirilgan — 1-qoida buzildi"
    assert barchasi[0].is_archived is True


async def test_test_arxivlanadi(
    client: AsyncClient, world: dict, session: AsyncSession
) -> None:
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)

    r = await client.post(f"/api/v1/tests/{t['id']}/archive", headers=_auth(token))
    assert r.status_code == 204

    barchasi = (await session.execute(select(Test))).scalars().all()
    assert len(barchasi) == 1
    assert barchasi[0].is_archived is True

    r = await client.get("/api/v1/tests", headers=_auth(token))
    assert r.json() == []


async def test_begona_ustoz_savollarni_kora_olmaydi(client: AsyncClient, world: dict) -> None:
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)
    await _add_question(client, token, t["id"])

    begona = await _token(client, "ts.begona")
    r = await client.get(f"/api/v1/tests/{t['id']}/questions", headers=_auth(begona))
    assert r.status_code == 403, r.text


# ─────────── TOʻGʻRI JAVOB SIZIB CHIQMASLIGI (1-qoida) ───────────


async def test_ustoz_togri_javobni_koradi(client: AsyncClient, world: dict) -> None:
    token = await _token(client, "ts.ustoz")
    t = await _create_test(client, token, world)
    await _add_question(client, token, t["id"], correct=1)

    r = await client.get(f"/api/v1/tests/{t['id']}/questions", headers=_auth(token))
    assert r.status_code == 200
    variantlar = r.json()[0]["options"]
    assert [o["is_correct"] for o in variantlar] == [False, True, False]


async def test_oquvchiga_togri_javob_yuborilmaydi(client: AsyncClient, world: dict) -> None:
    """Eng muhim test: javob sizib chiqsa butun modul maʼnosini yoʻqotadi."""
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    await _add_question(client, ustoz, t["id"], correct=1)
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
    )
    assert r.status_code == 200, r.text

    xom = r.text
    assert "is_correct" not in xom, "toʻgʻri javob javobda chiqib ketdi"
    for savol in r.json()["questions"]:
        for variant in savol["options"]:
            assert set(variant) == {"id", "text"}


async def test_oquvchi_ustoz_endpointiga_kira_olmaydi(client: AsyncClient, world: dict) -> None:
    """Ota-ona `/questions` orqali javoblarni olishga urinadi."""
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    r = await client.get(f"/api/v1/tests/{t['id']}/questions", headers=_auth(ota))
    assert r.status_code == 403, r.text


# ─────────────────────── Test ishlash ───────────────────────


async def test_ball_serverda_hisoblanadi(client: AsyncClient, world: dict) -> None:
    """TST-04: frontend ball yubormaydi."""
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"], correct=0, points=3)
    await _publish(client, ustoz, t["id"])

    togri = next(o["id"] for o in q["options"] if o["is_correct"])

    ota = await _token(client, "ts.ali")
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
    )
    attempt_id = r.json()["attempt_id"]

    r = await client.post(
        f"/api/v1/tests/attempts/{attempt_id}/submit",
        headers=_auth(ota),
        json={"answers": [{"question_id": q["id"], "selected": [togri]}]},
    )
    assert r.status_code == 200, r.text
    assert r.json()["score"] == 3
    assert r.json()["max_score"] == 3
    assert r.json()["percent"] == 100.0


async def test_notogri_javobga_ball_berilmaydi(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"], correct=0, points=2)
    await _publish(client, ustoz, t["id"])

    notogri = next(o["id"] for o in q["options"] if not o["is_correct"])

    ota = await _token(client, "ts.ali")
    attempt = (
        await client.post(
            f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
        )
    ).json()

    r = await client.post(
        f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit",
        headers=_auth(ota),
        json={"answers": [{"question_id": q["id"], "selected": [notogri]}]},
    )
    assert r.json()["score"] == 0
    assert r.json()["percent"] == 0.0


async def test_javobsiz_savol_ball_bermaydi(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"], points=5)
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    attempt = (
        await client.post(
            f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
        )
    ).json()

    r = await client.post(
        f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit",
        headers=_auth(ota),
        json={"answers": []},
    )
    assert r.status_code == 200, r.text
    assert r.json()["score"] == 0
    assert r.json()["max_score"] == 5
    assert q["points"] == 5


async def test_urinishlar_soni_cheklanadi(client: AsyncClient, world: dict) -> None:
    """TST-03: serverda, frontenddagi tugmani yashirish himoya emas."""
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world, attempts_allowed=1)
    q = await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    url = f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start"

    attempt = (await client.post(url, headers=_auth(ota))).json()
    await client.post(
        f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit",
        headers=_auth(ota),
        json={"answers": [{"question_id": q["id"], "selected": []}]},
    )

    r = await client.post(url, headers=_auth(ota))
    assert r.status_code == 409, r.text
    assert "Urinishlar tugadi" in r.json()["message"]


async def test_tugallanmagan_urinish_davom_etadi(client: AsyncClient, world: dict) -> None:
    """Sahifa yangilanganda urinish sarflanib ketmasin."""
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world, attempts_allowed=1)
    await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    url = f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start"

    birinchi = (await client.post(url, headers=_auth(ota))).json()
    ikkinchi = (await client.post(url, headers=_auth(ota))).json()
    assert birinchi["attempt_id"] == ikkinchi["attempt_id"]


async def test_ikki_marta_yuborilmaydi(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    attempt = (
        await client.post(
            f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
        )
    ).json()
    body = {"answers": [{"question_id": q["id"], "selected": []}]}
    url = f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit"

    assert (await client.post(url, headers=_auth(ota), json=body)).status_code == 200
    r = await client.post(url, headers=_auth(ota), json=body)
    assert r.status_code == 409, r.text


async def test_qoralama_test_ishlanmaydi(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    await _add_question(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
    )
    assert r.status_code == 409, r.text


async def test_hali_ochilmagan_test(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(
        client,
        ustoz,
        world,
        opens_at=(utcnow() + timedelta(days=1)).isoformat(),
        closes_at=(utcnow() + timedelta(days=2)).isoformat(),
    )
    await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
    )
    assert r.status_code == 409, r.text
    assert "ochilmagan" in r.json()["message"]


# ─────────────────────── Kirish nazorati (X-1) ───────────────────────


async def test_boshqa_sinf_oquvchisi_testni_ocha_olmaydi(
    client: AsyncClient, world: dict
) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota_b = await _token(client, "ts.vali")
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['vali'].id}/start", headers=_auth(ota_b)
    )
    assert r.status_code == 403, r.text


async def test_begona_otaona_boshqa_bolani_sorasa_403(
    client: AsyncClient, world: dict
) -> None:
    """URL dagi `student_id` ni oʻzgartirish."""
    ota_b = await _token(client, "ts.vali")
    r = await client.get(
        f"/api/v1/tests/students/{world['ali'].id}/available", headers=_auth(ota_b)
    )
    assert r.status_code == 403, r.text


async def test_ochiq_testlar_faqat_oz_sinfiniki(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota_a = await _token(client, "ts.ota_a")
    r = await client.get(
        f"/api/v1/tests/students/{world['ali'].id}/available", headers=_auth(ota_a)
    )
    assert r.status_code == 200
    assert [x["title"] for x in r.json()] == ["Kvadrat tenglamalar"]

    # 8-B oʻquvchisiga bu test koʻrinmaydi.
    ota_b = await _token(client, "ts.vali")
    r = await client.get(
        f"/api/v1/tests/students/{world['vali'].id}/available", headers=_auth(ota_b)
    )
    assert r.json() == []


async def test_begona_urinishni_yuborib_bolmaydi(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ali_t = await _token(client, "ts.ali")
    attempt = (
        await client.post(
            f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ali_t)
        )
    ).json()

    # Begona oʻquvchi urinishni yakunlay olmaydi.
    ota_b = await _token(client, "ts.vali")
    r = await client.post(
        f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit",
        headers=_auth(ota_b),
        json={"answers": [{"question_id": q["id"], "selected": []}]},
    )
    assert r.status_code == 403, r.text


# ─────────────────────── Natijalar (TST-05) ───────────────────────


async def test_natijalar_ustozga_korinadi(
    client: AsyncClient, world: dict, session: AsyncSession
) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"], correct=0, points=4)
    await _publish(client, ustoz, t["id"])

    togri = next(o["id"] for o in q["options"] if o["is_correct"])
    ota = await _token(client, "ts.ali")
    attempt = (
        await client.post(
            f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
        )
    ).json()
    await client.post(
        f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit",
        headers=_auth(ota),
        json={"answers": [{"question_id": q["id"], "selected": [togri]}]},
    )

    r = await client.get(f"/api/v1/tests/{t['id']}/results", headers=_auth(ustoz))
    assert r.status_code == 200, r.text
    assert len(r.json()) == 1
    assert r.json()[0]["full_name"] == "Abdullayev Ali"
    assert r.json()[0]["percent"] == 100.0

    # Roʻyxatdagi oʻrtacha ham yangilanadi.
    r = await client.get("/api/v1/tests", headers=_auth(ustoz))
    assert r.json()[0]["submitted_count"] == 1
    assert r.json()[0]["average_percent"] == 100.0

    saqlangan = (await session.execute(select(TestAttempt))).scalars().all()
    assert len(saqlangan) == 1
    assert saqlangan[0].score == 4


async def test_oquvchi_oz_natijalarini_koradi(client: AsyncClient, world: dict) -> None:
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    q = await _add_question(client, ustoz, t["id"])
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ali")
    attempt = (
        await client.post(
            f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
        )
    ).json()
    await client.post(
        f"/api/v1/tests/attempts/{attempt['attempt_id']}/submit",
        headers=_auth(ota),
        json={"answers": [{"question_id": q["id"], "selected": []}]},
    )

    r = await client.get(
        f"/api/v1/tests/students/{world['ali'].id}/attempts", headers=_auth(ota)
    )
    assert r.status_code == 200
    assert len(r.json()) == 1


async def test_ota_ona_farzand_nomidan_test_yecha_olmaydi(
    client: AsyncClient, world: dict
) -> None:
    """K6: koʻrish huquqi ≠ bajarish huquqi — vasiy testni boshlay olmaydi."""
    ustoz = await _token(client, "ts.ustoz")
    t = await _create_test(client, ustoz, world)
    await _add_question(client, ustoz, t["id"], correct=1)
    await _publish(client, ustoz, t["id"])

    ota = await _token(client, "ts.ota_a")  # haqiqiy vasiy, lekin oʻquvchi emas
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ota)
    )
    assert r.status_code == 403, r.text

    # Ustoz ham oʻquvchi nomidan boshlay olmaydi.
    r = await client.post(
        f"/api/v1/tests/{t['id']}/students/{world['ali'].id}/start", headers=_auth(ustoz)
    )
    assert r.status_code == 403, r.text


# ─────────────── Savollarni Excel'dan import (TST-06) ───────────────


def _savollar_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Savollar"
    ws.append([
        "Savol", "Ball",
        "Variant 1", "Variant 2", "Variant 3",
        "Variant 4", "Variant 5", "Variant 6",
    ])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _import_savollar(client: AsyncClient, token: str, test_id: str, rows: list[list]):
    return await client.post(
        f"/api/v1/tests/{test_id}/questions/import",
        headers=_auth(token),
        files={
            "file": (
                "savollar.xlsx",
                _savollar_xlsx(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


async def test_savollar_shabloni_yuklanadi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "ts.ustoz")
    r = await client.get("/api/v1/tests/questions/template", headers=_auth(t))
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "Savollar" in wb.sheetnames
    assert "Yoʻriqnoma" in wb.sheetnames


async def test_savollar_import_qilinadi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)

    r = await _import_savollar(
        client,
        t,
        test["id"],
        [
            ["Poytaxt qaysi shahar?", 1, "+ Toshkent", "Samarqand", "Buxoro", "", "", ""],
            ["2 + 2 = ?", 2, "+ 4", "5", "3", "", "", ""],
        ],
    )
    assert r.status_code == 201, r.text
    assert r.json()["added"] == 2

    savollar = await client.get(
        f"/api/v1/tests/{test['id']}/questions", headers=_auth(t)
    )
    rows = savollar.json()
    assert len(rows) == 2
    assert rows[0]["points"] == 1
    assert rows[1]["points"] == 2


async def test_tur_togri_javoblar_sonidan_aniqlanadi(
    client: AsyncClient, world: dict
) -> None:
    """Foydalanuvchi turni yozmaydi — ziddiyat chiqmasin."""
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)

    await _import_savollar(
        client,
        t,
        test["id"],
        [
            ["Bitta javob", 1, "+ A", "B", "C", "", "", ""],
            ["Bir nechta javob", 1, "+ A", "+ B", "C", "", "", ""],
        ],
    )
    rows = (
        await client.get(f"/api/v1/tests/{test['id']}/questions", headers=_auth(t))
    ).json()
    assert rows[0]["kind"] == "single"
    assert rows[1]["kind"] == "multiple"


async def test_buzuq_qator_importni_toxtatmaydi(
    client: AsyncClient, world: dict
) -> None:
    """60 ta savolli fayldagi bitta xato butun ishni yoʻqqa chiqarmasin."""
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)

    r = await _import_savollar(
        client,
        t,
        test["id"],
        [
            ["Yaxshi savol", 1, "+ A", "B", "", "", "", ""],
            ["Toʻgʻri javobsiz", 1, "A", "B", "", "", "", ""],
            ["Bitta variant", 1, "+ A", "", "", "", "", ""],
            ["Hammasi toʻgʻri", 1, "+ A", "+ B", "", "", "", ""],
            ["", 1, "+ A", "B", "", "", "", ""],
        ],
    )
    assert r.status_code == 201, r.text
    assert r.json()["added"] == 1
    # Har tashlangan qator sababi bilan qaytadi.
    assert len(r.json()["warnings"]) == 4


async def test_yaroqli_savol_yoq_bolsa_xato(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)

    r = await _import_savollar(
        client, t, test["id"], [["Toʻgʻri javobsiz", 1, "A", "B", "", "", "", ""]]
    )
    assert r.status_code == 422


async def test_savollar_mavjudlariga_qoshiladi(client: AsyncClient, world: dict) -> None:
    """Almashtirish EMAS: «hammasini oʻchirib qayta yozish» qaytarilmas yoʻqotish."""
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)
    await _add_question(client, t, test["id"])

    await _import_savollar(
        client, t, test["id"], [["Yangi savol", 1, "+ A", "B", "", "", "", ""]]
    )
    rows = (
        await client.get(f"/api/v1/tests/{test['id']}/questions", headers=_auth(t))
    ).json()
    assert len(rows) == 2


async def test_elon_qilingan_testga_import_qilinmaydi(
    client: AsyncClient, world: dict
) -> None:
    """`add_question` bilan bir xil qoida — oʻquvchi ishlayotgan test oʻzgarmasin."""
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)
    await _add_question(client, t, test["id"])
    await _publish(client, t, test["id"])

    r = await _import_savollar(
        client, t, test["id"], [["Yangi savol", 1, "+ A", "B", "", "", "", ""]]
    )
    assert r.status_code in (409, 422)


async def test_begona_ustoz_import_qila_olmaydi(client: AsyncClient, world: dict) -> None:
    t = await _token(client, "ts.ustoz")
    test = await _create_test(client, t, world)

    begona = await _token(client, "ts.begona")
    r = await _import_savollar(
        client, begona, test["id"], [["Savol", 1, "+ A", "B", "", "", "", ""]]
    )
    assert r.status_code == 403
