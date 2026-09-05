"""Rahbariyat endpointlari va rol darvozasi (DIR-01…DIR-04).

Eng muhim tekshiruv — OXIRGI ikkitasi: ustoz va ota-ona rahbariyat
hisobotiga kira olmasligi kerak. CLAUDE.md 7-qoida: rol tekshiruvi
serverda, frontendda yashirish himoya emas. Shu sabab har bir yangi
endpoint uchun «notoʻgʻri rol» testi yoziladi.
"""

import io
import uuid
from datetime import date, time, timedelta

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.core.timeutil import combine_local, local_today, utcnow
from app.models import (
    AcademicYear,
    AttendanceRecord,
    AttendanceStatus,
    AuditLog,
    Grade,
    GradeKind,
    Lesson,
    Permission,
    Role,
    RoleName,
    SchoolClass,
    Student,
    Subject,
    User,
    UserPermission,
)

PASSWORD = "Sinov12345!"  # noqa: S106
# Davr endi «bugun»dan orqaga sanaladi (audit Y5) — sana nisbiy.
DAY = local_today() - timedelta(days=1)


async def _roles(session: AsyncSession) -> dict[str, Role]:
    """Rollar migratsiyada seed qilingan — bu yerda faqat oʻqiladi.

    Avval har test oʻzi yaratardi; endi `t003` migratsiyasi ularni qatʼiy
    UUID bilan qoʻyadi va qayta yaratish `unique` cheklovini buzadi.
    """
    out = {r.name: r for r in (await session.execute(select(Role))).scalars()}
    yetishmaydi = {n.value for n in RoleName} - set(out)
    assert not yetishmaydi, f"Migratsiyada yoʻq rollar: {sorted(yetishmaydi)}"
    return out


async def _user(session: AsyncSession, roles: dict[str, Role], role: str, login: str) -> User:
    user = User(
        login=login,
        password_hash=hash_password(PASSWORD),
        last_name="Sinov",
        first_name=role.capitalize(),
    )
    user.roles = [roles[role]]
    session.add(user)
    await session.flush()
    return user


@pytest.fixture
async def school(session: AsyncSession) -> dict[str, object]:
    """Bitta sinf, bitta dars, ikkita oʻquvchi va toʻrtta rol."""
    roles = await _roles(session)

    director = await _user(session, roles, RoleName.DIRECTOR.value, "sinov.director")
    teacher = await _user(session, roles, RoleName.TEACHER.value, "sinov.teacher")
    parent = await _user(session, roles, RoleName.PARENT.value, "sinov.parent")
    academic = await _user(session, roles, RoleName.ACADEMIC.value, "sinov.academic")

    year = AcademicYear(
        name="2026-2027", starts_on=date(2026, 8, 24), ends_on=date(2027, 5, 25), is_current=True
    )
    session.add(year)
    await session.flush()

    subject = Subject(name="Matematika", short_name="Mat")
    session.add(subject)
    school_class = SchoolClass(academic_year_id=year.id, name="8-A", homeroom_teacher_id=teacher.id)
    session.add(school_class)
    await session.flush()

    students = [
        Student(class_id=school_class.id, last_name="Aliyev", first_name="Ali"),
        Student(class_id=school_class.id, last_name="Valiyev", first_name="Vali"),
    ]
    session.add_all(students)
    await session.flush()

    lesson = Lesson(
        class_id=school_class.id,
        subject_id=subject.id,
        teacher_id=teacher.id,
        lesson_date=DAY,
        period=1,
        room="8-A",
        starts_at=combine_local(DAY, time(8, 30)),
        ends_at=combine_local(DAY, time(9, 15)),
    )
    session.add(lesson)
    await session.flush()

    # Biri keldi, biri kelmadi → davomat aynan 50% boʻlishi kerak.
    session.add(
        AttendanceRecord(
            lesson_id=lesson.id,
            student_id=students[0].id,
            status=AttendanceStatus.PRESENT.value,
            marked_by_id=teacher.id,
            marked_at=lesson.ends_at,
        )
    )
    session.add(
        AttendanceRecord(
            lesson_id=lesson.id,
            student_id=students[1].id,
            status=AttendanceStatus.ABSENT.value,
            marked_by_id=teacher.id,
            marked_at=lesson.ends_at,
        )
    )
    session.add(
        Grade(
            student_id=students[0].id,
            subject_id=subject.id,
            lesson_id=lesson.id,
            teacher_id=teacher.id,
            kind=GradeKind.CURRENT.value,
            value=4,
            max_value=5,
            weight=1,
        )
    )
    await session.flush()

    return {
        "director": director,
        "teacher": teacher,
        "parent": parent,
        "academic": academic,
    }


async def _token(client: AsyncClient, login: str) -> str:
    resp = await client.post("/api/v1/auth/login", json={"login": login, "password": PASSWORD})
    assert resp.status_code == 200, resp.text
    return resp.json()["access_token"]


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def test_login_returns_user_and_sets_refresh_cookie(
    client: AsyncClient, school: dict[str, object]
) -> None:
    resp = await client.post(
        "/api/v1/auth/login", json={"login": "sinov.director", "password": PASSWORD}
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["user"]["roles"] == [RoleName.DIRECTOR.value]
    # Refresh token javob tanasida QAYTMAYDI — faqat cookie'da.
    assert "refresh_token" not in body
    assert "tarbion_rt" in resp.cookies


async def test_login_rejects_wrong_password(client: AsyncClient, school: dict[str, object]) -> None:
    resp = await client.post(
        "/api/v1/auth/login", json={"login": "sinov.director", "password": "xato"}
    )
    assert resp.status_code == 401


async def test_overview_counts_come_from_database(
    client: AsyncClient, school: dict[str, object]
) -> None:
    token = await _token(client, "sinov.director")
    resp = await client.get("/api/v1/director/overview", headers=_auth(token))
    assert resp.status_code == 200
    body = resp.json()

    assert body["total_students"] == 2
    assert body["total_classes"] == 1
    assert body["total_teachers"] == 1
    assert body["lessons_planned"] == 1
    # Ikki yozuvdan biri «keldi» → 50%.
    assert body["attendance_percent"] == 50.0
    assert body["average_grade"] == 4.0
    assert body["attendance_trend"] == [{"date": DAY.isoformat(), "percent": 50.0}]


async def test_overview_period_window_excludes_older_lessons(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """`days` oynasi BUGUNDAN orqaga sanaladi (audit Y5).

    Oynadan tashqaridagi eski dars hisobga kirmaydi.
    """
    lesson = (await session.execute(select(Lesson))).scalars().first()
    assert lesson is not None

    old_day = DAY - timedelta(days=40)
    session.add(
        Lesson(
            class_id=lesson.class_id,
            subject_id=lesson.subject_id,
            teacher_id=lesson.teacher_id,
            lesson_date=old_day,
            period=2,
            room="8-A",
            starts_at=combine_local(old_day, time(9, 25)),
            ends_at=combine_local(old_day, time(10, 10)),
        )
    )
    await session.flush()

    token = await _token(client, "sinov.director")

    wide = await client.get("/api/v1/director/overview", params={"days": 60}, headers=_auth(token))
    assert wide.json()["lessons_planned"] == 2

    narrow = await client.get("/api/v1/director/overview", params={"days": 7}, headers=_auth(token))
    assert narrow.json()["lessons_planned"] == 1


async def test_classes_row_has_attendance_and_homeroom(
    client: AsyncClient, school: dict[str, object]
) -> None:
    token = await _token(client, "sinov.director")
    resp = await client.get("/api/v1/director/classes", headers=_auth(token))
    assert resp.status_code == 200
    rows = resp.json()
    assert len(rows) == 1
    row = rows[0]
    assert row["name"] == "8-A"
    assert row["student_count"] == 2
    assert row["attendance_percent"] == 50.0
    assert row["average_grade"] == 4.0
    assert row["homeroom_teacher_name"] == "Sinov Teacher"


async def test_teachers_lists_only_staff_with_lessons(
    client: AsyncClient, school: dict[str, object]
) -> None:
    """Direktor ham, ota-ona ham darsi yoʻq — roʻyxatda faqat ustoz chiqadi."""
    token = await _token(client, "sinov.director")
    resp = await client.get("/api/v1/director/teachers", headers=_auth(token))
    assert resp.status_code == 200
    rows = resp.json()
    assert len(rows) == 1
    assert rows[0]["homeroom_class_name"] == "8-A"
    assert rows[0]["lessons_planned"] == 1
    assert rows[0]["grades_given"] == 1
    assert rows[0]["average_grade_given"] == 4.0


@pytest.mark.parametrize(
    "path",
    ["/api/v1/director/overview", "/api/v1/director/classes", "/api/v1/director/teachers"],
)
async def test_director_endpoints_require_authentication(
    client: AsyncClient, school: dict[str, object], path: str
) -> None:
    assert (await client.get(path)).status_code == 401
    bad = await client.get(path, headers=_auth("soxta.token.qiymat"))
    assert bad.status_code == 401


@pytest.mark.parametrize("login", ["sinov.teacher", "sinov.parent"])
async def test_teacher_and_parent_cannot_read_director_reports(
    client: AsyncClient, school: dict[str, object], login: str
) -> None:
    """Eng muhim test: notoʻgʻri rol 403 oladi, 200 emas."""
    token = await _token(client, login)
    resp = await client.get("/api/v1/director/overview", headers=_auth(token))
    assert resp.status_code == 403


async def test_academic_head_can_read_director_reports(
    client: AsyncClient, school: dict[str, object]
) -> None:
    """Oʻquv boʻlimi imtihon va ustozlar faoliyatini barcha sinflar
    kesimida koʻradi — shu sabab hisobotlarga kirishi kerak."""
    token = await _token(client, "sinov.academic")
    resp = await client.get("/api/v1/director/classes", headers=_auth(token))
    assert resp.status_code == 200


async def test_me_returns_current_user(client: AsyncClient, school: dict[str, object]) -> None:
    token = await _token(client, "sinov.director")
    resp = await client.get("/api/v1/auth/me", headers=_auth(token))
    assert resp.status_code == 200
    assert resp.json()["roles"] == [RoleName.DIRECTOR.value]
    uuid.UUID(resp.json()["id"])


# ─────────── Faoliyat koʻrsatkichlari (DIR-04, 2026-09-03 soʻrovi) ───────────


async def test_teachers_faoliyat_korsatkichlari(
    client: AsyncClient, school: dict[str, object]
) -> None:
    """Nol koʻrsatkich ham QAYTADI — interfeys uni izoh bilan koʻrsatadi.

    Ilgari bu maydonlar umuman yoʻq edi va rahbar «ustoz imtihon
    oldimi, vazifa berdimi» degan savolga javob topa olmasdi.
    """
    token = await _token(client, "sinov.director")
    rows = (
        await client.get("/api/v1/director/teachers", headers=_auth(token))
    ).json()
    ustoz = rows[0]

    # Fiksturada imtihon ham, vazifa ham yoʻq — nol boʻlishi KERAK,
    # maydon yoʻqolib ketishi emas.
    assert ustoz["exams_held"] == 0
    assert ustoz["homework_given"] == 0

    # Bitta dars, oʻsha darsda davomat belgilangan.
    assert ustoz["lessons_planned"] == 1
    assert ustoz["lessons_with_attendance"] == 1


async def test_davomat_belgilangan_dars_bir_marta_sanaladi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Bitta darsda 25 ta davomat yozuvi boʻladi — dars BIR marta sanalsin."""
    token = await _token(client, "sinov.director")
    rows = (
        await client.get("/api/v1/director/teachers", headers=_auth(token))
    ).json()
    # Fiksturada bitta darsda IKKITA davomat yozuvi bor.
    assert rows[0]["lessons_with_attendance"] == 1


async def test_ortacha_ball_davr_ichida(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Oʻrtacha ball ham tanlangan davrga boʻysunadi.

    Ilgari bu koʻrsatkichda sana filtri yoʻq edi: rahbar 7 kunni
    tanlaganda davomat oʻzgarardi, oʻrtacha ball esa butun tarix
    boʻyicha qotib turardi — ikki koʻrsatkich yonma-yon turib turli
    davrni bildirardi.
    """
    from sqlalchemy import update

    from app.models import Grade

    token = await _token(client, "sinov.director")

    # Bahoni 60 kun orqaga suramiz.
    await session.execute(
        update(Grade).values(created_at=utcnow() - timedelta(days=60))
    )
    await session.flush()

    qisqa = (
        await client.get("/api/v1/director/overview?days=7", headers=_auth(token))
    ).json()
    uzun = (
        await client.get("/api/v1/director/overview?days=90", headers=_auth(token))
    ).json()

    assert qisqa["average_grade"] == 0.0, "eski baho qisqa davrga tushmasligi kerak"
    assert uzun["average_grade"] == 4.0, "uzun davrda esa hisobga olinsin"


# ──────── «Oʻtilgan dars» va davr qamrovi (loyiha egasining soʻrovi) ────────
#
# Uchta jimgina xato tuzatildi (2026-09-03):
#   · ustozlar soni davr filtriga boʻysunmasdi;
#   · ustozning «oʻtilgan darslari» butun yilga oldindan generatsiya
#     qilingan darslarni, jumladan KELAJAKDAGILARNI ham sanardi;
#   · davomat foizi nechta yozuvdan hisoblangani koʻrinmasdi.


async def _kelajak_darsi(session: AsyncSession, school: dict, *, kun_keyin: int) -> Lesson:
    """Jadval butun yilga generatsiya qilinadi — kelajakdagi dars odatiy holat."""
    ustoz = school["teacher"]
    sinf = (await session.execute(select(SchoolClass))).scalars().first()
    fan = (await session.execute(select(Subject))).scalars().first()
    kun = local_today() + timedelta(days=kun_keyin)
    lesson = Lesson(
        class_id=sinf.id,
        subject_id=fan.id,
        teacher_id=ustoz.id,
        lesson_date=kun,
        period=2,
        starts_at=combine_local(kun, time(8, 30)),
        ends_at=combine_local(kun, time(9, 15)),
    )
    session.add(lesson)
    await session.flush()
    return lesson


async def test_ustoz_hisobi_davrga_boysunadi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Ustozlar soni ham tanlangan davr ichida sanaladi.

    Fiksturadagi yagona dars kecha boʻlgan. 1 kunlik davr — faqat
    bugun — ichida darsi bor ustoz yoʻq.
    """
    token = await _token(client, "sinov.director")

    keng = await client.get(
        "/api/v1/director/overview", params={"days": 30}, headers=_auth(token)
    )
    assert keng.json()["total_teachers"] == 1

    tor = await client.get("/api/v1/director/overview", params={"days": 1}, headers=_auth(token))
    assert tor.json()["total_teachers"] == 0


async def test_overviewda_otilgan_va_rejadagi_dars_ajratiladi(
    client: AsyncClient, school: dict[str, object]
) -> None:
    token = await _token(client, "sinov.director")
    body = (await client.get("/api/v1/director/overview", headers=_auth(token))).json()

    assert body["lessons_planned"] == 1
    # Davomat belgilangan — dars haqiqatan oʻtilgani shundan bilinadi.
    assert body["lessons_with_attendance"] == 1
    # Foiz ikkita yozuvdan hisoblangan; rahbar buni koʻrib turishi kerak.
    assert body["attendance_records"] == 2


async def test_kelajakdagi_dars_ustoz_hisobiga_kirmaydi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Jadval butun yilga tuziladi — «oʻtilgan dars» ni oshirib yubormasin."""
    token = await _token(client, "sinov.director")
    oldin = (await client.get("/api/v1/director/teachers", headers=_auth(token))).json()
    assert oldin[0]["lessons_planned"] == 1

    await _kelajak_darsi(session, school, kun_keyin=7)

    keyin = (await client.get("/api/v1/director/teachers", headers=_auth(token))).json()
    assert keyin[0]["lessons_planned"] == 1, "kelajakdagi dars sanalib ketdi"


async def test_faqat_kelajakdagi_darsi_bor_ustoz_royxatdan_yoqolmaydi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Roʻyxat jadvaldan, koʻrsatkich esa oʻtgan darslardan.

    Aks holda dushanbadan boshlanadigan ustoz yakshanba kuni
    ustozlar roʻyxatidan butunlay yoʻqolardi.
    """
    roles = await _roles(session)
    yangi = await _user(session, roles, RoleName.TEACHER.value, "sinov.teacher2")
    sinf = (await session.execute(select(SchoolClass))).scalars().first()
    fan = (await session.execute(select(Subject))).scalars().first()
    kun = local_today() + timedelta(days=3)
    session.add(
        Lesson(
            class_id=sinf.id,
            subject_id=fan.id,
            teacher_id=yangi.id,
            lesson_date=kun,
            period=3,
            starts_at=combine_local(kun, time(10, 0)),
            ends_at=combine_local(kun, time(10, 45)),
        )
    )
    await session.flush()

    token = await _token(client, "sinov.director")
    rows = (await client.get("/api/v1/director/teachers", headers=_auth(token))).json()
    qator = next(r for r in rows if r["id"] == str(yangi.id))
    assert qator["lessons_planned"] == 0


async def test_sinf_davomati_nechta_yozuvdan_chiqqani_koresatiladi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Yozuv boʻlmasa «0%» va «hali belgilanmagan» bir xil koʻrinmasin.

    Ilgari ikkala holat ham qizil «0%» boʻlib chiqardi va rahbar buni
    «sinf darsga kelmayapti» deb oʻqirdi.
    """
    token = await _token(client, "sinov.director")
    rows = (await client.get("/api/v1/director/classes", headers=_auth(token))).json()
    assert rows[0]["attendance_records"] == 2

    yangi_sinf = SchoolClass(
        academic_year_id=(await session.execute(select(AcademicYear))).scalars().first().id,
        name="9-B",
    )
    session.add(yangi_sinf)
    await session.flush()

    rows = (await client.get("/api/v1/director/classes", headers=_auth(token))).json()
    bosh = next(r for r in rows if r["name"] == "9-B")
    assert bosh["attendance_percent"] == 0
    assert bosh["attendance_records"] == 0



# ─────────────── Hisobotlarni eksport (DIR-08, X-13) ───────────────


async def _huquq(session: AsyncSession, user, permission: Permission) -> None:
    """Huquqni toʻgʻridan-toʻgʻri beradi — `grant` chaqiruvchida
    `permissions.grant` talab qiladi va bu testda ortiqcha qadam."""
    session.add(UserPermission(user_id=user.id, permission=permission.value))
    await session.commit()


async def test_huquqsiz_direktor_eksport_qila_olmaydi(
    client: AsyncClient, school: dict[str, object]
) -> None:
    """Rol yetarli emas — yuklab olish alohida beriladigan huquq."""
    token = await _token(client, "sinov.director")
    r = await client.get(
        "/api/v1/director/reports/sinflar/export",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 403


async def test_sinflar_hisoboti_excelga_chiqadi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    await _huquq(session, school["director"], Permission.REPORTS_EXPORT)

    token = await _token(client, "sinov.director")
    r = await client.get(
        "/api/v1/director/reports/sinflar/export",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 200, r.text
    assert r.headers["cache-control"] == "private, no-store"
    assert ".xlsx" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    # Sana faylning ICHIDA turadi — «bu qaysi kundagi holat?» savoli
    # bir oydan keyin ham javobsiz qolmasin.
    assert "Maʼlumot" in wb.sheetnames
    varaq = wb["Sinflar kesimi"]
    assert varaq["A1"].value == "Sinf"
    assert varaq.freeze_panes == "A2"


async def test_ustozlar_hisoboti_excelga_chiqadi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    await _huquq(session, school["director"], Permission.REPORTS_EXPORT)

    token = await _token(client, "sinov.director")
    r = await client.get(
        "/api/v1/director/reports/ustozlar/export",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert "Ustozlar faoliyati" in wb.sheetnames


async def test_nomalum_hisobot_turi_422(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    await _huquq(session, school["director"], Permission.REPORTS_EXPORT)

    token = await _token(client, "sinov.director")
    r = await client.get(
        "/api/v1/director/reports/allaqanday/export",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 422


async def test_eksport_auditga_tushadi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """X-13: eng ehtimolli sizib chiqish hujum emas, xodim."""
    await _huquq(session, school["director"], Permission.REPORTS_EXPORT)

    token = await _token(client, "sinov.director")
    await client.get(
        "/api/v1/director/reports/sinflar/export",
        headers={"Authorization": f"Bearer {token}"},
    )

    rows = (
        (
            await session.execute(
                select(AuditLog).where(AuditLog.object_type == "report_export")
            )
        )
        .scalars()
        .all()
    )
    assert len(rows) == 1
    assert rows[0].new_value["kind"] == "sinflar"
    assert rows[0].actor_id == school["director"].id


async def test_oquv_bolimi_qarzdorlikni_eksport_qila_olmaydi(
    client: AsyncClient, school: dict[str, object], session: AsyncSession
) -> None:
    """Oʻquv boʻlimi hisobotlarni koʻradi, lekin moliyani KOʻRMAYDI.

    `reports.export` huquqi berilgan boʻlsa ham — moliya tekshiruvi
    alohida turadi (`assert_finance_admin`).
    """
    await _huquq(session, school["academic"], Permission.REPORTS_EXPORT)

    token = await _token(client, "sinov.academic")
    r = await client.get(
        "/api/v1/director/reports/qarzdorlik/export",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 403

    # Sinflar kesimi esa ochiq.
    ok = await client.get(
        "/api/v1/director/reports/sinflar/export",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert ok.status_code == 200
